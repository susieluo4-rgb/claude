#!/usr/bin/env python3
"""
润铭 组合持仓自动同步 — 从 Zejun Outlook 邮件 → 润铭.md

数据流：
  Zejun邮件 (Outlook 收件箱)
    → 附件 润铭portfolio YYYYMMDD.xlsx
    → Sheet1: 包含 Ticker, Name, 润铭%, Price, Chg_Pct, contri润铭, sector
    → 更新 润铭.md 的持仓明细表
    → 同时更新 润铭_组合视图.md 的统计数据

用法：
  python3 runming_update.py --dry-run           # 查看变更（不修改文件）
  python3 runming_update.py --auto              # 实际执行更新
  python3 runming_update.py --from-xlsx /tmp/portfolio.xlsx  # 从本地 XLSX 预览
"""
import argparse
import json
import os
import re
import sys
import base64
import tempfile
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

# Outlook 认证
OUTLOOK_SCRIPT = str(Path.home() / ".claude" / "skills" / "outlook-microsoft" / "scripts")
sys.path.insert(0, OUTLOOK_SCRIPT)
from outlook_auth import get_valid_token, GRAPH_BASE_URL

import requests

# ========== 路径配置 ==========
RUNMING_MD_PATH = Path.home() / "Research" / "Vault_基金经理Agent" / "润铭.md"
RUNMING_VIEW_PATH = Path.home() / "Research" / "Vault_共享知识库" / "wiki" / "portfolio" / "润铭_组合视图.md"

# ========== 1. 从 Outlook 获取最新 XLSX ==========

def api_call(method, path, token, data=None, params=None):
    """调用 Microsoft Graph API"""
    url = f"{GRAPH_BASE_URL}/v1.0{path}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        resp = requests.request(method, url, headers=headers, json=data, params=params, timeout=30)
    except requests.exceptions.RequestException as e:
        return None, f"网络错误: {e}"
    if resp.status_code == 401:
        return None, "Token 已失效，请运行 outlook_auth.py refresh"
    if not resp.ok:
        try:
            err = resp.json()
            return None, f"API 错误: {err.get('error', {}).get('message', resp.text)[:200]}"
        except Exception:
            return None, f"HTTP {resp.status_code}"
    return resp.json() if resp.text else {}, ""


def download_attachment(path, token):
    """下载附件（二进制数据）"""
    url = f"{GRAPH_BASE_URL}/v1.0{path}"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        resp = requests.get(url, headers=headers, timeout=30)
    except requests.exceptions.RequestException as e:
        return None, f"网络错误: {e}"
    if resp.status_code == 401:
        return None, "Token 已失效，请运行 outlook_auth.py refresh"
    if not resp.ok:
        return None, f"HTTP {resp.status_code}"
    return resp.content, ""


def fetch_latest_xlsx(token):
    """搜索最新 portfolio 邮件（来自 zejun.ma@binyuancapital.com）并下载 XLSX 附件"""
    print("搜索 Zejun 的 portfolio 邮件...")

    # 获取最近的邮件
    params = {
        "$top": 20,
        "$orderby": "receivedDateTime desc",
        "$select": "id,subject,from,receivedDateTime,hasAttachments"
    }
    data, err = api_call("GET", "/me/messages", token, params=params)
    if err:
        print(f"  搜索失败: {err}")
        return None, err

    msgs = data.get("value", [])
    target = None
    for msg in msgs:
        subj = msg.get("subject", "")
        sender = msg.get("from", {}).get("emailAddress", {}).get("address", "")
        if "portfolio" in subj.lower() and "zejun.ma" in sender.lower():
            target = msg
            break

    if not target:
        return None, f"未找到来自 zejun.ma@binyuancapital.com 的 portfolio 邮件"

    msg_id = target.get("id")
    subj = target.get("subject", "")
    print(f"  ✅ 找到: {subj}")

    # 获取邮件详情和附件
    msg_data, err = api_call("GET", f"/me/messages/{msg_id}", token, params={"$expand": "attachments"})
    if err:
        return None, f"获取邮件详情失败: {err}"

    attachments = msg_data.get("attachments", [])
    target_att = None

    # 查找 "润铭portfolio" Excel 附件
    for att in attachments:
        name = att.get("name", "")
        if "润铭portfolio" in name and (name.endswith(".xlsx") or name.endswith(".xls")):
            target_att = att
            break

    if not target_att:
        return None, f"邮件中未找到 '润铭portfolio' Excel 附件，仅找到: {[a.get('name') for a in attachments]}"

    att_id = target_att.get("id")
    att_name = target_att.get("name")
    print(f"  ✅ 找到附件: {att_name}")

    # 下载附件
    xlsx_bytes, err = download_attachment(f"/me/messages/{msg_id}/attachments/{att_id}/$value", token)
    if err:
        return None, f"下载附件失败: {err}"

    # 写入临时文件
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(xlsx_bytes)
        temp_path = f.name

    return temp_path, ""


# ========== 2. 解析 XLSX 数据 ==========

def parse_runming_xlsx(xlsx_path):
    """解析润铭 portfolio Excel 文件"""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active  # Sheet1
    except Exception as e:
        return None, f"打开 Excel 文件失败: {e}"

    holdings = []
    sector_stats = defaultdict(float)
    market_stats = {"A股": 0, "港股": 0, "北交所": 0}

    # 获取表头行（第一行）
    headers = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers = list(row)
        break

    # 固定列索引（基于 Excel 的实际结构）
    # 第 0 列：日期
    # 第 1 列：Ticker
    # 第 2 列：Name
    # 第 3 列：润铭
    # 第 8 列：sector
    ticker_col = 1
    name_col = 2
    weight_col = 3
    sector_col = 8

    # 验证列
    if headers[ticker_col] != "Ticker" or headers[weight_col] != "润铭":
        return None, f"Excel 列结构不符合预期。Ticker at {ticker_col}: {headers[ticker_col]}, Weight at {weight_col}: {headers[weight_col]}"

    # 逐行读取数据
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[ticker_col]:  # 跳过空行
            continue

        ticker = str(row[ticker_col]).strip()
        name = str(row[name_col]).strip() if row[name_col] else "Unknown"
        weight_raw = row[weight_col] if row[weight_col] else 0
        sector = str(row[sector_col]).strip() if row[sector_col] else "其他"

        # 验证 ticker 格式（应该包含 . 和市场标识）
        valid_markets = ['.SZ', '.SH', '.HK', '.BJ']
        is_valid_ticker = '.' in ticker and any(ticker.upper().endswith(m.upper()) for m in valid_markets)
        if not is_valid_ticker:
            continue

        # 处理权重值
        if isinstance(weight_raw, str):
            # 如果是百分比字符串，提取数字
            weight_str = weight_raw.replace('%', '').strip()
            try:
                weight_float = float(weight_str) / 100
                weight_pct = f"{weight_float*100:.2f}%"
            except ValueError:
                weight_float = 0
                weight_pct = "0.00%"
        elif isinstance(weight_raw, (int, float)):
            weight_float = weight_raw
            if weight_raw < 1:  # 小数形式
                weight_pct = f"{weight_raw*100:.2f}%"
            else:  # 百分比形式
                weight_pct = f"{weight_raw:.2f}%"
        else:
            weight_float = 0
            weight_pct = "0.00%"

        # ticker 保持原样，不做修改
        ticker_clean = ticker

        holdings.append({
            "ticker": ticker_clean,
            "name": name,
            "weight": weight_pct,
            "sector": sector,
            "weight_float": weight_float
        })

        # 统计行业和市场
        sector_stats[sector] += weight_float

        # 判断市场
        if ticker.endswith(".BJ"):
            market_stats["北交所"] += weight_float
        elif ticker.endswith(".HK"):
            market_stats["港股"] += weight_float
        elif ticker.endswith((".SZ", ".SH")):
            market_stats["A股"] += weight_float

    # 按权重排序
    holdings.sort(key=lambda x: x["weight_float"], reverse=True)

    return {
        "holdings": holdings,
        "sector_stats": dict(sector_stats),
        "market_stats": market_stats,
        "count": len(holdings),
        "total_weight": sum(h["weight_float"] for h in holdings)
    }, ""


# ========== 3. 读取旧持仓 & 生成对比 ==========

def parse_old_holdings(md_path):
    """从现有的 Markdown 文件提取旧持仓数据"""
    if not Path(md_path).exists():
        return {}

    old_holdings = {}
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                # 匹配表格行：| ticker | name | weight% |
                if '|' not in line:
                    continue
                parts = [p.strip() for p in line.split('|')]
                if len(parts) >= 4 and parts[0] == '' and parts[1] and parts[2] and '%' in parts[3]:
                    # 跳过表头和分隔符行
                    if parts[1] in ('Ticker', '---') or '---' in parts[3]:
                        continue
                    ticker = parts[1]
                    weight_str = parts[3].replace('%', '').strip()
                    try:
                        weight = float(weight_str) / 100
                        old_holdings[ticker] = weight
                    except ValueError:
                        pass
    except Exception:
        pass

    return old_holdings


def generate_comparison_md(old_holdings, new_holdings):
    """生成持仓对比报告"""
    md = f"# 润铭组合持仓对比 — {datetime.now().strftime('%Y-%m-%d')}\n\n"

    # 构建新持仓字典
    new_dict = {h["ticker"]: h["weight_float"] for h in new_holdings}

    # 分类变化
    new = {}  # 新增
    removed = {}  # 清空
    changed = {}  # 变化

    for ticker in set(list(old_holdings.keys()) + list(new_dict.keys())):
        old_w = old_holdings.get(ticker, 0)
        new_w = new_dict.get(ticker, 0)

        if old_w == 0 and new_w > 0:
            new[ticker] = new_w
        elif old_w > 0 and new_w == 0:
            removed[ticker] = old_w
        elif old_w > 0 and new_w > 0 and abs(old_w - new_w) > 0.0001:
            changed[ticker] = (old_w, new_w)

    # 统计变化
    md += f"## 变化汇总\n\n"
    md += f"| 类型 | 数量 |\n"
    md += f"|------|------|\n"
    md += f"| 新增标的 | {len(new)} |\n"
    md += f"| 清仓标的 | {len(removed)} |\n"
    md += f"| 权重调整 | {len(changed)} |\n\n"

    # 新增标的
    if new:
        md += f"## 新增标的 ({len(new)})\n\n"
        md += "| Ticker | Name | 权重 |\n"
        md += "|--------|------|------|\n"
        new_list = [(t, n) for t, n in sorted([(ticker, next((h["name"] for h in new_holdings if h["ticker"] == ticker), "")) for ticker in new], key=lambda x: new[x[0]], reverse=True)]
        for ticker, name in new_list:
            display_ticker = ticker[1:] if ticker.startswith("0") and ticker.endswith(".HK") else ticker
            weight_pct = new[ticker] * 100
            md += f"| {display_ticker} | {name} | {weight_pct:.2f}% |\n"
        md += "\n"

    # 清仓标的
    if removed:
        md += f"## 清仓标的 ({len(removed)})\n\n"
        md += "| Ticker | 原权重 |\n"
        md += "|--------|--------|\n"
        for ticker in sorted(removed.keys(), key=lambda x: removed[x], reverse=True):
            display_ticker = ticker[1:] if ticker.startswith("0") and ticker.endswith(".HK") else ticker
            md += f"| {display_ticker} | {removed[ticker]*100:.2f}% |\n"
        md += "\n"

    # 权重调整
    if changed:
        md += f"## 权重调整 ({len(changed)})\n\n"
        md += "| Ticker | 旧权重 | 新权重 | 变化 |\n"
        md += "|--------|--------|--------|------|\n"
        changed_sorted = sorted(changed.items(), key=lambda x: abs(x[1][1] - x[1][0]), reverse=True)
        for ticker, (old_w, new_w) in changed_sorted:
            delta = new_w - old_w
            sign = "+" if delta > 0 else ""
            display_ticker = ticker[1:] if ticker.startswith("0") and ticker.endswith(".HK") else ticker
            md += f"| {display_ticker} | {old_w*100:.2f}% | {new_w*100:.2f}% | {sign}{delta*100:.2f}% |\n"
        md += "\n"

    return md


# ========== 4. 生成 Markdown ==========

def generate_runming_md(data):
    """生成润铭.md 的新内容"""
    holdings = data["holdings"]

    md = "# 组合持仓 润铭\n\n"
    md += "## 持仓明细\n\n"
    md += "| Ticker    | Name      | 润铭     |\n"
    md += "| --------- | --------- | -------- |\n"

    for h in holdings:
        # 格式化 ticker（处理港股的前导 0）
        ticker = h["ticker"]
        if ticker.startswith("0") and ticker.endswith(".HK"):
            ticker = ticker[1:]  # 去掉前导 0

        md += f"| {ticker} | {h['name']} | {h['weight']} |\n"

    md += "\n"

    return md


def generate_runming_view_md(data):
    """生成润铭_组合视图.md"""
    holdings = data["holdings"]
    market_stats = data["market_stats"]
    sector_stats = data["sector_stats"]

    md = "# 润铭_组合视图\n\n"
    md += f"> 最后更新：{datetime.now().strftime('%Y-%m-%d')}\n"
    md += "> 数据来源：[[Vault_基金经理Agent/润铭]]（用户主维护源）\n"
    md += f"> 持仓数量：{data['count']} 条持仓记录 / {len(set(h['ticker'] for h in holdings))} 家独特公司\n\n"

    md += "---\n\n"
    md += "## 组合概览\n\n"
    md += "| 市场 | 数量 | 权重合计 | 代表性标的 |\n"
    md += "|------|------|---------|----------|\n"

    # 统计各市场的持仓数量
    for market in ["A股", "港股", "北交所"]:
        market_holdings = [h for h in holdings if (
            (market == "A股" and h["ticker"].endswith((".SZ", ".SH"))) or
            (market == "港股" and h["ticker"].endswith(".HK")) or
            (market == "北交所" and h["ticker"].endswith(".BJ"))
        )]

        count = len(market_holdings)
        weight = market_stats.get(market, 0)
        top3 = ", ".join([h["name"] for h in market_holdings[:3]])

        md += f"| {market} | {count} | {weight*100:.1f}% | {top3} |\n"

    md += "\n*注：权重合计可能超过 100% 因为市场分类重复或四舍五入。*\n\n"
    md += "---\n\n"
    md += "## 完整持仓明细\n\n"
    md += "| 代码 | 公司名称 | 权重 |\n"
    md += "|------|---------|------|\n"

    for h in holdings:
        ticker = h["ticker"]
        if ticker.startswith("0") and ticker.endswith(".HK"):
            ticker = ticker[1:]
        md += f"| {ticker} | {h['name']} | {h['weight']} |\n"

    md += "\n"

    return md


# ========== 5. 主函数 ==========

def main():
    parser = argparse.ArgumentParser(description="润铭组合持仓自动同步")
    parser.add_argument("--dry-run", action="store_true", help="仅查看变更，不修改文件")
    parser.add_argument("--auto", action="store_true", help="实际执行更新")
    parser.add_argument("--from-xlsx", type=str, help="从本地 XLSX 文件读取数据")

    args = parser.parse_args()

    print("=" * 60)
    print("润铭组合持仓自动同步")
    print("=" * 60)

    # Step 1: 获取 XLSX
    if args.from_xlsx:
        xlsx_path = args.from_xlsx
        print(f"使用本地文件: {xlsx_path}")
    else:
        print("\n[1/3] 从 Outlook 获取最新 XLSX...")
        token, _, _ = get_valid_token()
        if not token:
            print(f"❌ 认证失败")
            sys.exit(1)

        xlsx_path, err = fetch_latest_xlsx(token)
        if err:
            print(f"❌ 获取失败: {err}")
            sys.exit(1)

    # Step 2: 解析数据
    print("\n[2/3] 解析 Excel 数据...")
    data, err = parse_runming_xlsx(xlsx_path)
    if err:
        print(f"❌ 解析失败: {err}")
        sys.exit(1)

    print(f"✅ 成功解析: {data['count']} 个持仓")
    print(f"   总权重: {data['total_weight']*100:.1f}%")
    print(f"   A股: {len([h for h in data['holdings'] if h['ticker'].endswith(('.SZ', '.SH'))])}")
    print(f"   港股: {len([h for h in data['holdings'] if h['ticker'].endswith('.HK')])}")
    print(f"   北交所: {len([h for h in data['holdings'] if h['ticker'].endswith('.BJ')])}")

    # Step 3: 生成新 Markdown
    print("\n[3/3] 生成 Markdown...")
    new_runming_md = generate_runming_md(data)
    new_runming_view_md = generate_runming_view_md(data)

    # Step 4: 显示预览或执行更新
    if args.dry_run or not args.auto:
        print("\n" + "=" * 60)
        print("📄 润铭.md 预览（前 30 行）")
        print("=" * 60)
        print("\n".join(new_runming_md.split("\n")[:30]))
        print("\n...")

        if args.dry_run:
            print("\n✅ 预览完成（使用 --auto 执行实际更新）")
        else:
            response = input("\n执行更新? (y/n): ")
            if response.lower() != 'y':
                print("已取消")
                sys.exit(0)

    if args.auto or (not args.dry_run and input("\n执行更新? (y/n): ").lower() == 'y'):
        # 生成对比报告（必须在更新文件之前）
        old_holdings = parse_old_holdings(RUNMING_MD_PATH)
        comparison_md = generate_comparison_md(old_holdings, data["holdings"])

        # 更新文件
        with open(RUNMING_MD_PATH, 'w', encoding='utf-8') as f:
            f.write(new_runming_md)
        print(f"✅ 已更新: {RUNMING_MD_PATH}")

        with open(RUNMING_VIEW_PATH, 'w', encoding='utf-8') as f:
            f.write(new_runming_view_md)
        print(f"✅ 已更新: {RUNMING_VIEW_PATH}")

        # 保存对比报告
        reports_dir = Path.home() / "0.Agent 投研总监" / "reports"
        reports_dir.mkdir(parents=True, exist_ok=True)
        comparison_path = reports_dir / f"润铭对比_{datetime.now().strftime('%Y%m%d')}.md"
        with open(comparison_path, 'w', encoding='utf-8') as f:
            f.write(comparison_md)
        print(f"✅ 已生成: {comparison_path}")

        print("\n✅ 同步完成！")


if __name__ == "__main__":
    main()
