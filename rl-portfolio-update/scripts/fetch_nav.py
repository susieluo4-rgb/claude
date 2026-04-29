#!/usr/bin/env python3
"""
每日净值自动拉取 — 从 Outlook 邮件（润铭 xlsx）+ Hereford Funds 网页（CIF）
获取最新净值，自动更新 Wiki 组合视图

属于 rl-portfolio-update skill 的净值组件。

数据源：
  - 润铭 → Ma, ZeJun <zejun.ma@binyuancapital.com> "portfolio YYYYMMDD"
           附件：绝对收益净值表 YYYYMMDD.xlsx（Sheet"净值", C列=润铭NAV）
  - CIF  → https://herefordfunds.com/funds/bin-yuan-china-innovation-fund

用法:
  # 增量更新（默认）— 只抓取最新一天数据
  python3 fetch_nav.py [runming|cif|both]    # 默认 both

  # 回填完整历史（首次使用或重建时）
  python3 fetch_nav.py --backfill runming    # 从 xlsx 回填润铭全部历史净值
"""
import sys
import json
import os
import re
from datetime import datetime, date
from pathlib import Path

# ── 路径配置 ──
SCRIPT_DIR = Path(__file__).parent
OUTLOOK_AUTH = SCRIPT_DIR.parents[1] / "outlook-microsoft" / "scripts"
sys.path.insert(0, str(OUTLOOK_AUTH))
from outlook_auth import get_valid_token, GRAPH_BASE_URL

WIKI_DIR = Path.home() / "Research" / "Vault_共享知识库" / "wiki" / "portfolio"
STATE_FILE = SCRIPT_DIR / "nav_fetch_state.json"

import requests
import openpyxl
import csv
import io


# ── 状态管理 ──
def load_state():
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            return json.load(f)
    return {"runming_last_date": "", "cif_last_date": "", "last_check": ""}

def save_state(state):
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2, ensure_ascii=False)


# ── Outlook API 工具 ──
def api_get(path, params=None):
    token, _, _ = get_valid_token()
    if not token:
        return None
    url = f"{GRAPH_BASE_URL}/v1.0{path}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    resp = requests.get(url, headers=headers, params=params, timeout=60)
    if not resp.ok:
        print(f"  API 错误: {resp.status_code} {resp.text[:200]}")
        return None
    return resp.json()


def download_attachment(msg_id, att_name, save_path):
    """下载邮件附件到本地"""
    # 1. 获取附件列表
    data = api_get(f"/me/messages/{msg_id}/attachments")
    if not data:
        return False
    for a in data.get("value", []):
        if a.get("name") == att_name:
            att_id = a["id"]
            # 下载附件内容
            token, _, _ = get_valid_token()
            if not token:
                return False
            url = f"{GRAPH_BASE_URL}/v1.0/me/messages/{msg_id}/attachments/{att_id}/$value"
            headers = {"Authorization": f"Bearer {token}"}
            resp = requests.get(url, headers=headers, timeout=60, stream=True)
            if not resp.ok:
                print(f"  下载失败: {resp.status_code}")
                return False
            save_path.parent.mkdir(parents=True, exist_ok=True)
            with open(save_path, "wb") as f:
                for chunk in resp.iter_content(8192):
                    f.write(chunk)
            return True
    print(f"  附件 {att_name} 未找到")
    return False


# ── 润铭 NAV 拉取 ──

def _find_runming_xlsx():
    """搜索最新 ZeJun 邮件，下载 xlsx 到 /tmp，返回 (tmp_path, date_str) 或 (None, None)"""
    data = api_get("/me/messages", {
        "$top": 10,
        "$search": '"zejun.ma@binyuancapital.com"',
        "$select": "id,subject,from,receivedDateTime,hasAttachments"
    })
    if not data:
        return None, None

    for msg in data.get("value", []):
        subj = msg.get("subject", "")
        sender = msg.get("from", {}).get("emailAddress", {}).get("address", "")
        if "zejun.ma" not in sender:
            continue
        if not msg.get("hasAttachments"):
            continue

        msg_id = msg["id"]
        rcv = msg["receivedDateTime"][:10]
        print(f"  找到邮件: {subj} ({rcv}) sender={sender}")

        att_name = None
        att_data = api_get(f"/me/messages/{msg_id}/attachments")
        if att_data:
            for a in att_data.get("value", []):
                if "绝对收益净值表" in a.get("name", ""):
                    att_name = a["name"]
                    break
            if not att_name:
                for a in att_data.get("value", []):
                    if a.get("name", "").endswith(".xlsx"):
                        att_name = a["name"]
                        break

        if not att_name:
            print("  未找到 xlsx 附件")
            continue

        tmp_path = Path("/tmp") / att_name
        print(f"  下载附件: {att_name}")
        if download_attachment(msg_id, att_name, tmp_path):
            return tmp_path, rcv
        else:
            return None, None

    print("  ❌ 未找到润铭 NAV 邮件")
    return None, None


def fetch_runming_nav():
    """从 Outlook 获取润铭最新净值（增量更新，只取最后一行）"""
    print("=" * 50)
    print("📌 润铭 NAV 拉取")

    tmp_path, rcv = _find_runming_xlsx()
    if not tmp_path:
        return None

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb["净值"]

        last_date = None
        last_nav = None
        for row in range(ws.max_row, 1, -1):
            c_val = ws.cell(row=row, column=3).value
            a_val = ws.cell(row=row, column=1).value
            if c_val is not None and isinstance(c_val, (int, float)) and a_val is not None:
                last_date = a_val
                last_nav = c_val
                break

        if last_date and last_nav:
            if isinstance(last_date, datetime):
                d = last_date.strftime("%Y-%m-%d")
            else:
                d = str(last_date)[:10]
            print(f"  最新净值: {d} → {last_nav:.4f}")
            return (d, round(last_nav, 4))
        else:
            print("  未找到有效净值数据")
            return None
    except Exception as e:
        print(f"  解析失败: {e}")
        return None
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def backfill_runming_nav():
    """全量回填：从 xlsx 提取所有历史净值，重写润铭_组合视图.md 净值序列表格"""
    print("=" * 50)
    print("📌 润铭历史 NAV 回填")

    tmp_path, rcv = _find_runming_xlsx()
    if not tmp_path:
        return False

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb["净值"]

        rows = []
        for row in range(2, ws.max_row + 1):
            a_val = ws.cell(row=row, column=1).value
            c_val = ws.cell(row=row, column=3).value
            if c_val is not None and isinstance(c_val, (int, float)) and a_val is not None:
                if isinstance(a_val, datetime):
                    d = a_val.strftime("%Y-%m-%d")
                else:
                    d = str(a_val)[:10]
                rows.append((d, round(c_val, 4)))

        rows.sort(key=lambda x: x[0])
        print(f"  共 {len(rows)} 行历史数据")
        print(f"  范围: {rows[0][0]} → {rows[-1][0]}")

        return _write_runming_full_table(rows)
    except Exception as e:
        print(f"  解析失败: {e}")
        return False
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def _write_runming_full_table(rows):
    """用完整净值序列重写 润铭_组合视图.md 的 ## 净值序列与归因分析 章节"""
    path = WIKI_DIR / "润铭_组合视图.md"
    if not path.exists():
        print(f"  ❌ 文件不存在: {path}")
        return False

    with open(path, encoding="utf-8") as f:
        content = f.read()

    first_date = rows[0][0]
    last_date = rows[-1][0]

    # 构建新表格行
    table_rows = "\n".join(
        f"| {d} | {nav:.4f} | {nav:.4f} |" for d, nav in rows
    )

    new_section = f"""## 净值序列与归因分析

> 数据来源：润铭绝对收益净值表（Ma ZeJun 邮件）
> 最后更新：{last_date}

| 日期 | 单位净值 | 累计净值 |
|------|---------|---------|
{table_rows}

### 分析要点

- 数据从 **{first_date}** 开始记录（日频，自动更新）
- 单位净值 = 润铭基金实际净值（2017年10月成立，基值≈1.0）
- 累计净值 = 单位净值（此处为同一值，历史累计收益待补充）
"""

    # 替换现有章节或追加
    section_pattern = r"## 净值序列与归因分析[\s\S]*?(?=\n## |\Z)"
    if re.search(section_pattern, content):
        new_content = re.sub(section_pattern, new_section, content)
    else:
        # 追加到最后
        new_content = content.rstrip() + "\n\n" + new_section

    with open(path, "w", encoding="utf-8") as f:
        f.write(new_content)

    print(f"  ✅ 已回填 {len(rows)} 行净值数据到 {path.name}")
    print(f"  范围: {first_date} → {last_date}")
    return True


# ── CIF NAV 拉取 ──
CIF_NAV_URL = "https://herefordfunds.com/funds/bin-yuan-china-innovation-fund"

def fetch_cif_nav():
    """从 Hereford Funds 公开网页获取 CIF 最新净值"""
    print("=" * 50)
    print("📌 CIF NAV 拉取")
    print(f"  来源: {CIF_NAV_URL}")

    nav_value = None
    nav_date = None

    try:
        resp = requests.get(CIF_NAV_URL, timeout=30)
        resp.raise_for_status()
        html = resp.text

        # 查找净值表格: <th>NAV</th> <th>Daily %</th> <th>Date</th>
        # 定位到 FI share class 行
        # 结构: <tr><td>FI</td><td>USD</td><td>NAV</td><td>Daily%</td><td class="date">Date</td>...
        import re

        # 匹配 FI share class 的整行
        # 示例: <td>FI</td>...<td>USD</td><td>100.47 </td><td class="text-success">0.75%</td><td class="2026-04-27 00:00:00">27/04/2026</td>
        pattern = r'<td>FI</td>\s*<td>USD</td>\s*<td>([\d.]+)\s*</td>\s*<td[^>]*>[^<]*</td>\s*<td class="([\d-]+)\s'
        match = re.search(pattern, html)

        if match:
            nav_value = float(match.group(1))
            nav_date_str = match.group(2)
            nav_date = nav_date_str[:10]
            print(f"  最新净值: {nav_date} → {nav_value:.2f} (FI USD)")
        else:
            # 兜底: 搜所有出现的 NAV 数值
            nav_pattern = r'<td>([\d.]+)\s*</td>\s*<td class="text-\w+">([-\d.]+)%</td>\s*<td class="([\d-]+)'
            all_navs = re.findall(nav_pattern, html)
            if all_navs:
                # 第一个出现的往往是 FI share class
                nav_value = float(all_navs[0][0])
                nav_date_str = all_navs[0][2]
                nav_date = nav_date_str[:10]
                print(f"  最新净值(兜底): {nav_date} → {nav_value:.2f}")
            else:
                print("  ❌ 未找到 NAV 数据")
                return None

    except requests.RequestException as e:
        print(f"  ❌ 请求失败: {e}")
        return None
    except Exception as e:
        print(f"  ❌ 解析失败: {e}")
        return None

    if nav_value and nav_date:
        return (nav_date, round(nav_value, 2))
    return None


# ── Wiki 更新 ──
def update_runming_wiki(date_str, nav):
    """更新 润铭_组合视图.md 净值序列"""
    path = WIKI_DIR / "润铭_组合视图.md"
    if not path.exists():
        print(f"  ❌ 文件不存在: {path}")
        return False

    with open(path, encoding="utf-8") as f:
        content = f.read()

    # 检查是否已有该日期
    if f"| {date_str} |" in content:
        print(f"  ⏭️  日期 {date_str} 已存在，跳过")
        return True

    # 查找已有的净值表格最后一行
    # 格式: | 日期 | 单位净值 | 累计净值 |
    # build_benchmark_html.py 需要3列: date, unit_nav, cum_nav
    # 此处 unit_nav 和 cum_nav 都是原始NAV值，方便后续计算
    pattern = r"\|\s*(\d{4}-\d{2}-\d{2})\s*\|\s*([\d.]+)\s*\|\s*([\d.]+)\s*\|"
    matches = re.findall(pattern, content)

    # 累计净值 = 单位净值（相对基准由 build_benchmark_html.py 计算）
    new_row = f"| {date_str} | {nav:.4f} | {nav:.4f} |"

    # 检查是否存在 ## 净值序列 章节
    if "## 净值序列" in content or "## 净值序列与归因分析" in content:
        # 追加到最后一行表格后
        table_pattern = r"(\| 日期 \| 单位净值 \|.*?\|[\s\S]*?)(?=\n\n|\Z)"
        # 更精确：在表格末尾追加
        lines = content.split("\n")
        # 找到最后一个以 | 开头的行（表格行）
        last_table_idx = None
        for i in range(len(lines) - 1, -1, -1):
            if lines[i].strip().startswith("|") and "--" not in lines[i]:
                last_table_idx = i
                break

        if last_table_idx is not None:
            # 插入新行
            lines.insert(last_table_idx + 1, new_row)
            new_content = "\n".join(lines)
            with open(path, "w", encoding="utf-8") as f:
                f.write(new_content)
            print(f"  ✅ 已追加 {date_str} → {nav:.2f}")
            return True
        else:
            print("  ⚠️  找到了章节但未找到表格行")
            return False
    else:
        # 没有净值章节，创建并更新最后更新日期
        print("  ⚠️  润铭_组合视图.md 缺少净值序列章节，需要手动补充全量历史数据")
        # 至少追加一个单行
        # 用 raw NAV 值（此处 unit_nav = cum_nav = raw nav）
        nav_row = f"| {date_str} | {nav:.4f} | {nav:.4f} |"
        section = f"""
## 净值序列与归因分析

> 数据来源：润铭绝对收益净值表（Ma ZeJun 邮件）
> 最后更新：{date_str}

| 日期 | 单位净值 | 累计净值 |
|------|---------|---------|
{nav_row}

### 分析要点

- 数据从 **{date_str}** 开始记录（日频，自动更新）
- 单位净值 = 润铭基金实际净值（2017年10月成立，基值≈1.0）
- 累计净值 = 单位净值（此处为同一值，历史累计收益待补充）
"""
        content += section
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"  ✅ 已创建净值序列章节，首行 {date_str} → {nav:.2f}")
        return True


def update_cif_wiki(date_str, nav):
    """更新 CIF_组合视图.md 净值序列"""
    path = WIKI_DIR / "CIF_组合视图.md"
    if not path.exists():
        print(f"  ❌ 文件不存在: {path}")
        return False

    with open(path, encoding="utf-8") as f:
        content = f.read()

    if f"| {date_str} |" in content:
        print(f"  ⏭️  日期 {date_str} 已存在，跳过")
        return True

    # CIF 格式: | 日期 | 单位净值 | 相对期初涨跌 |
    pattern = r"\|\s*(\d{4}-\d{2}-\d{2})\s*\|\s*([\d.]+)\s*\|"
    matches = re.findall(pattern, content)

    prev_nav = None
    first_nav = None
    if matches:
        prev_nav = float(matches[-1][1])
        first_nav = float(matches[0][1])

    cum_pct = None
    if first_nav:
        cum_pct = (nav - first_nav) / first_nav * 100

    pct_str = "—（期初）"
    if prev_nav:
        pct = (nav - prev_nav) / prev_nav * 100
        if cum_pct is not None:
            pct_str = f"{pct:+.2f}%"
            cum_str = f"{cum_pct:+.2f}%"
        else:
            pct_str = f"{pct:+.2f}%"
    else:
        cum_str = "—（期初）"

    if cum_pct is not None:
        new_row = f"| {date_str} | {nav:.2f} | {cum_str} |"
    else:
        new_row = f"| {date_str} | {nav:.2f} | {pct_str} |"

    # 追加到表格末尾（在净值序列表格区域内）
    lines = content.split("\n")
    last_table_idx = None
    for i in range(len(lines) - 1, -1, -1):
        stripped = lines[i].strip()
        if stripped.startswith("|") and "--" not in stripped and "日期" not in stripped:
            last_table_idx = i
            break

    if last_table_idx is not None:
        lines.insert(last_table_idx + 1, new_row)
        new_content = "\n".join(lines)
        with open(path, "w", encoding="utf-8") as f:
            f.write(new_content)
        print(f"  ✅ CIF 已追加 {date_str} → {nav:.2f}")
        return True
    else:
        print("  ⚠️  找不到 CIF 净值表格位置")
        return False


# ── 主入口 ──
def main():
    import argparse
    parser = argparse.ArgumentParser(description="组合净值自动拉取")
    parser.add_argument("mode", nargs="?", default="both",
                        choices=["runming", "cif", "both"],
                        help="增量更新模式")
    parser.add_argument("--backfill", choices=["runming"],
                        help="全量历史回填（目前仅支持 runming）")
    args = parser.parse_args()

    state = load_state()

    if args.backfill == "runming":
        success = backfill_runming_nav()
        if success:
            state["runming_last_date"] = ""
        state["last_check"] = datetime.now().isoformat()
        save_state(state)
        return

    results = []

    if args.mode in ("runming", "both"):
        result = fetch_runming_nav()
        if result:
            d, nav = result
            state["runming_last_date"] = d
            update_runming_wiki(d, nav)
            results.append(f"润铭: {d} NAV={nav}")
        else:
            results.append("润铭: 未获取到")

    if args.mode in ("cif", "both"):
        result = fetch_cif_nav()
        if result:
            d, nav = result
            state["cif_last_date"] = d
            update_cif_wiki(d, nav)
            results.append(f"CIF: {d} NAV={nav}")
        else:
            results.append("CIF: 未获取到")

    state["last_check"] = datetime.now().isoformat()
    save_state(state)

    print("\n" + "=" * 50)
    print("📊 结果汇总:")
    for r in results:
        print(f"  • {r}")
    print("=" * 50)


if __name__ == "__main__":
    main()
