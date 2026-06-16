#!/usr/bin/env python3
"""
CIF 组合持仓自动同步 — 从 Sunny Outlook 邮件 → CIF.md

数据流：
  Sunny邮件 (Outlook 收件箱)
    → 附件 Portfolio H YYYYMMDD.xlsx
    → Sheet: HFHC (Innovation Fund = CIF)
    → 列: A=TICKER, B=TODAY'S WEIGHT, U=中文名
    → ticker_map.json 将 HFHC ticker 转为 CIF.md 格式
    → 更新 CIF.md + CIF_组合视图.md

用法：
  python3 cif_portfolio_sync.py --dry-run     # 查看变更（不修改文件）
  python3 cif_portfolio_sync.py --auto         # 实际执行更新
  python3 cif_portfolio_sync.py --from-xlsx /tmp/portfolio_h.xlsx  # 从本地 XLSX 预览
"""
import argparse
import json
import os
import re
import sys
import base64
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

# Outlook 认证
OUTLOOK_SCRIPT = str(Path.home() / ".claude" / "skills" / "outlook-microsoft" / "scripts")
sys.path.insert(0, OUTLOOK_SCRIPT)
from outlook_auth import get_valid_token, GRAPH_BASE_URL

import requests

# ========== 路径配置 ==========
SKILL_DIR = Path.home() / ".claude" / "skills" / "rl-portfolio-monitor"
TICKER_MAP_PATH = SKILL_DIR / "cif_name_ticker_map.json"
CIF_MD_PATH = Path.home() / "Research" / "Vault_基金经理Agent" / "CIF.md"
CIF_VIEW_PATH = Path.home() / "Research" / "Vault_共享知识库" / "wiki" / "portfolio" / "CIF_组合视图.md"

# ========== 1. 从 Outlook 获取最新 XLSX ==========

def api_call(method, path, token, data=None, params=None):
    """调用 Microsoft Graph API"""
    url = f"{GRAPH_BASE_URL}/v1.0{path}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        resp = requests.request(method, url, headers=headers, json=data, params=params, timeout=30)
    except requests.exceptions.RequestException as e:
        return None, f"网络错误: {e}"
    if resp.status_code == 401:
        return None, "Token 已失效，请运行 outlook_auth.py refresh"
    if not resp.ok:
        try:
            err = resp.json()
            return None, f"API 错误: {err.get('error', {}).get('message', resp.text)[:200]}"
        except Exception:
            return None, f"HTTP {resp.status_code}"
    return resp.json() if resp.text else {}, ""


def find_folder_by_path(token, folder_path):
    """根据路径查找文件夹（如 "2.Fund Related/CIF"）"""
    parts = folder_path.split("/")
    folder_id = "inbox"

    for part in parts:
        # 获取当前文件夹的子文件夹
        params = {"$select": "id,displayName"}
        data, err = api_call("GET", f"/me/mailFolders/{folder_id}/childFolders", token, params=params)
        if err:
            return None, f"获取子文件夹失败: {err}"

        # 查找匹配的子文件夹（忽略大小写）
        found = False
        part_lower = part.lower()
        for child in data.get("value", []):
            if child.get("displayName", "").lower() == part_lower:
                folder_id = child.get("id")
                found = True
                break

        if not found:
            return None, f"未找到文件夹: {part}"

    return folder_id, ""


def fetch_latest_xlsx(token):
    """搜索最新 Portfolio H 邮件并下载 XLSX 附件"""
    print("搜索 Sunny 的 Portfolio H 邮件...")

    # 查找 "2.Fund Related/CIF" 文件夹
    folder_id, err = find_folder_by_path(token, "2.Fund Related/CIF")
    if err:
        print(f"  查找文件夹失败: {err}，尝试搜索主收件箱...")
        folder_id = "inbox"
    else:
        print(f"  在 '2.Fund Related/CIF' 文件夹中搜索...")

    # 简单查询：只用 $top，不组合 $orderby 和其他复杂参数
    params = {
        "$top": 50,
        "$select": "id,subject,from,receivedDateTime,hasAttachments"
    }
    data, err = api_call("GET", f"/me/mailFolders/{folder_id}/messages", token, params=params)
    if err:
        print(f"  查询失败: {err}")
        return None, err

    msgs = data.get("value", [])
    target = None
    for msg in msgs:
        subj = msg.get("subject", "")
        sender = msg.get("from", {}).get("emailAddress", {}).get("address", "")
        if "Portfolio H" in subj and "sunny.wang" in sender.lower():
            target = msg
            break
    if not target:
        # 放宽条件：任何含 Portfolio H 的邮件
        for msg in msgs:
            if "Portfolio H" in msg.get("subject", ""):
                target = msg
                break
    if not target:
        return None, "未找到 Portfolio H 邮件"

    msg_id = target["id"]
    date = target.get("receivedDateTime", "")[:10]
    print(f"  找到: {target['subject']} ({date})")

    # 获取附件列表
    att_data, err = api_call("GET", f"/me/messages/{msg_id}/attachments", token)
    if err:
        return None, f"获取附件列表失败: {err}"

    xlsx_att = None
    for att in att_data.get("value", []):
        name = att.get("name", "")
        if "Portfolio H" in name and (name.endswith(".xlsx") or name.endswith(".xls")):
            xlsx_att = att
            break
    if not xlsx_att:
        return None, "未找到 Portfolio H XLSX 附件"

    print(f"  下载附件: {xlsx_att['name']}")

    # 下载附件内容
    att_id = xlsx_att["id"]
    # 如果是 fileAttachment，用 $value 端点
    att_type = xlsx_att.get("@odata.type", "")
    if "fileAttachment" in att_type:
        # 方法1: 直接读 contentBytes
        att_detail, err = api_call("GET", f"/me/messages/{msg_id}/attachments/{att_id}", token)
        if err:
            return None, f"下载附件失败: {err}"
        content_bytes = att_detail.get("contentBytes", "")
        if not content_bytes:
            return None, "附件内容为空"
        raw = base64.b64decode(content_bytes)
    else:
        # itemAttachment 或其它类型
        return None, f"不支持的附件类型: {att_type}"

    # 保存到临时文件
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="portfolio_h_")
    tmp.write(raw)
    tmp.close()
    print(f"  保存到: {tmp.name}")
    return tmp.name, target


# ========== 2. 解析 HFHC Sheet ==========

def parse_hfhc_sheet(xlsx_path):
    """解析 HFHC sheet，返回 {normalized_ticker: {name, weight}}"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "HFHC" not in wb.sheetnames:
        return None, f"HFHC sheet 不存在，可用: {wb.sheetnames}"

    ws = wb["HFHC"]
    holdings = {}
    for r in range(3, ws.max_row + 1):
        ticker = ws.cell(r, 1).value
        weight = ws.cell(r, 2).value
        name = ws.cell(r, 21).value
        if ticker and isinstance(weight, (int, float)):
            # 排除 NAV 行和现金
            name_str = str(name).strip() if name else ""
            if "NAV" in str(ticker).upper() or "NAV" in name_str:
                continue
            # 排除基金表现跟踪行和指数（不是实际持仓）
            ticker_str = str(ticker).strip().upper()
            if "INNOVATION FUND" in ticker_str or "MSCI CHINA" in ticker_str or "INDEX" in ticker_str:
                continue
            if weight > 0 and weight < 1:  # 真实持仓
                holdings[str(ticker).strip()] = {
                    "name": name_str,
                    "weight": weight
                }

    return holdings, ""


# ========== 3. Ticker 归一化 ==========

def load_ticker_map():
    """加载 HFHC → CIF.md ticker 映射"""
    if not TICKER_MAP_PATH.exists():
        print(f"错误: ticker_map 不存在，请先创建 {TICKER_MAP_PATH}")
        sys.exit(1)
    with open(TICKER_MAP_PATH) as f:
        return json.load(f)


def normalize_holdings(hfhc_holdings, ticker_map):
    """将 HFHC ticker 转为 CIF.md 格式，返回 {cif_ticker: {name, weight, is_new}}"""
    result = {}
    unmapped = []
    for hfhc_ticker, info in hfhc_holdings.items():
        if hfhc_ticker in ticker_map:
            cif_ticker = ticker_map[hfhc_ticker]
            result[cif_ticker] = {"name": info["name"], "weight": info["weight"], "is_new": False}
        else:
            unmapped.append((hfhc_ticker, info["name"], info["weight"]))

    return result, unmapped


# ========== 4. 读取当前 CIF.md ==========

def parse_cif_md():
    """解析 CIF.md，返回 {ticker: {name, weight_str, weight_pct}}"""
    if not CIF_MD_PATH.exists():
        return None, f"CIF.md 不存在: {CIF_MD_PATH}"

    text = CIF_MD_PATH.read_text(encoding="utf-8")
    holdings = {}
    table_pattern = re.compile(r'^\| (.+) \| (.+) \| (.+) \|$', re.MULTILINE)
    for m in table_pattern.finditer(text):
        ticker = m.group(1).strip()
        name = m.group(2).strip()
        weight_str = m.group(3).strip()
        if ticker and name and ticker != "Ticker" and "---" not in ticker:
            try:
                weight_pct = float(weight_str.replace("%", "").strip()) if weight_str != "-" else 0
            except ValueError:
                weight_pct = 0
            holdings[ticker] = {"name": name, "weight_str": weight_str, "weight_pct": weight_pct}
    return holdings, text


# ========== 5. 生成 Diff ==========

def compute_diff(cif_holdings, new_holdings, unmapped, xlsx_path, email_subject, email_date):
    """计算新旧持仓差异"""
    # 权重变化
    changes = []
    for ticker, info in new_holdings.items():
        new_pct = round(info["weight"] * 100, 2)
        if ticker in cif_holdings:
            old_pct = cif_holdings[ticker]["weight_pct"]
            diff = round(new_pct - old_pct, 2)
            if abs(diff) > 0.01:
                changes.append({
                    "ticker": ticker,
                    "name": info["name"],
                    "old": old_pct,
                    "new": new_pct,
                    "diff": diff,
                    "type": "update"
                })
        else:
            changes.append({
                "ticker": ticker,
                "name": info["name"],
                "old": 0,
                "new": new_pct,
                "diff": new_pct,
                "type": "new"
            })

    # 移除的持仓（在 CIF.md 中但不在 HFHC 中）
    for ticker, info in cif_holdings.items():
        if ticker not in new_holdings and ticker != "Cash" and info["name"] != "现金":
            changes.append({
                "ticker": ticker,
                "name": info["name"],
                "old": info["weight_pct"],
                "new": 0,
                "diff": -info["weight_pct"],
                "type": "removed"
            })

    return changes


def generate_comparison_md(changes, email_subject, email_date):
    """生成持仓对比 Markdown 报告"""
    updates = [c for c in changes if c["type"] == "update"]
    news = [c for c in changes if c["type"] == "new"]
    removed = [c for c in changes if c["type"] == "removed"]

    md = f"# CIF 组合持仓对比 — {email_date}\n\n"
    md += f"> 数据源：{email_subject}\n\n"

    # 统计
    md += f"## 变化汇总\n\n"
    md += f"| 类型 | 数量 |\n"
    md += f"|------|------|\n"
    md += f"| 权重调整 | {len(updates)} |\n"
    md += f"| 新增标的 | {len(news)} |\n"
    md += f"| 清仓标的 | {len(removed)} |\n\n"

    # 权重变化
    if updates:
        md += f"## 权重调整 ({len(updates)})\n\n"
        md += f"| Ticker | 名称 | 原权重 | 新权重 | 变化 |\n"
        md += f"|--------|------|--------|--------|------|\n"
        for c in sorted(updates, key=lambda x: abs(x["diff"]), reverse=True):
            arrow = "▲" if c["diff"] > 0 else "▼"
            md += f"| {c['ticker']} | {c['name']} | {c['old']:.2f}% | {c['new']:.2f}% | {arrow}{abs(c['diff']):.2f}% |\n"
        md += "\n"

    # 新增持仓
    if news:
        md += f"## 新增持仓 ({len(news)})\n\n"
        md += f"| Ticker | 名称 | 权重 |\n"
        md += f"|--------|------|------|\n"
        for c in sorted(news, key=lambda x: x["new"], reverse=True):
            md += f"| {c['ticker']} | {c['name']} | {c['new']:.2f}% |\n"
        md += "\n"

    # 清仓标的
    if removed:
        md += f"## 清仓标的 ({len(removed)})\n\n"
        md += f"| Ticker | 名称 | 原权重 |\n"
        md += f"|--------|------|--------|\n"
        for c in sorted(removed, key=lambda x: x["old"], reverse=True):
            md += f"| {c['ticker']} | {c['name']} | {c['old']:.2f}% |\n"
        md += "\n"

    return md


def print_diff_report(changes, unmapped, email_subject, email_date):
    """打印变更报告"""
    updates = [c for c in changes if c["type"] == "update"]
    news = [c for c in changes if c["type"] == "new"]
    removed = [c for c in changes if c["type"] == "removed"]

    print(f"\n{'='*60}")
    print(f"  CIF 持仓同步 — {email_date}")
    print(f"  来源: {email_subject}")
    print(f"{'='*60}")

    if updates:
        print(f"\n  ▶ 权重变化 ({len(updates)}):")
        print(f"  {'Ticker':20s} {'名称':12s} {'原权重':>8s} {'新权重':>8s} {'变化':>8s}")
        print(f"  {'-'*56}")
        for c in sorted(updates, key=lambda x: abs(x["diff"]), reverse=True):
            arrow = "▲" if c["diff"] > 0 else "▼"
            print(f"  {c['ticker']:20s} {c['name']:12s} {c['old']:7.2f}% {c['new']:7.2f}% {arrow}{abs(c['diff']):6.2f}%")

    if news:
        print(f"\n  ⚠ 新增持仓 ({len(news)}):")
        for c in sorted(news, key=lambda x: x["new"], reverse=True):
            print(f"    + {c['ticker']:20s} {c['name']:12s} {c['new']:6.2f}%")

    if removed:
        print(f"\n  ⚠ 可能已清仓 ({len(removed)}):")
        for c in sorted(removed, key=lambda x: x["old"], reverse=True):
            print(f"    - {c['ticker']:20s} {c['name']:12s} (原 {c['old']:5.2f}%)")

    if unmapped:
        print(f"\n  ⚠ 未映射持仓 (需更新 ticker_map.json):")
        for t, n, w in unmapped:
            print(f"    ? {t:30s} {n:12s} {w*100:6.2f}%")

    news_sum = sum(c["new"] for c in news)
    removed_sum = sum(c["old"] for c in removed)

    print(f"\n  ---")
    print(f"  总变化: {len(updates)} 更新 + {len(news)} 新增 - {len(removed)} 移除")
    print(f"  新增合计: {news_sum:.2f}% | 移除合计: {removed_sum:.2f}%")
    print(f"{'='*60}\n")


# ========== 6. 更新 CIF.md ==========

def update_cif_md(cif_text, new_holdings, removed_tickers):
    """更新 CIF.md 中的权重值"""
    lines = cif_text.split("\n")
    new_lines = []
    table_line_pattern = re.compile(r'^\|\s*([\w.]+?)\s*\|\s*(.+?)\s*\|\s*([\d.]+%)\s*\|$')

    for line in lines:
        m = table_line_pattern.match(line)
        if m:
            ticker = m.group(1)
            name = m.group(2)
            weight_str = m.group(3)

            if ticker in new_holdings:
                new_pct = round(new_holdings[ticker]["weight"] * 100, 2)
                new_weight = f"{new_pct}%"
                new_lines.append(line.replace(weight_str, new_weight))
            elif ticker in removed_tickers:
                new_lines.append(f"<!-- {line.strip()} -->")
            else:
                new_lines.append(line)
        else:
            new_lines.append(line)
    return "\n".join(new_lines)


def update_cif_view_md(date_str, holdings):
    """更新 CIF_组合视图.md 的日期和权重"""
    if not CIF_VIEW_PATH.exists():
        return None, "CIF_组合视图.md 不存在"

    text = CIF_VIEW_PATH.read_text(encoding="utf-8")

    # 更新 "最后更新" 日期
    date_pattern = re.compile(r'最后更新[：:]\s*\d{4}-\d{2}-\d{2}')
    if date_pattern.search(text):
        text = date_pattern.sub(f"最后更新：{date_str}", text)
    else:
        print("  警告: CIF_组合视图.md 中未找到 '最后更新' 日期")

    # 更新持仓表格权重列
    table_line_pattern = re.compile(r'^\|\s*([\w.]+?)\s*\|\s*(.+?)\s*\|\s*([\d.]+%)\s*\|')
    lines = text.split("\n")
    new_lines = []
    for line in lines:
        m = table_line_pattern.match(line)
        if m:
            ticker = m.group(1)
            weight_str = m.group(3)
            if ticker in holdings:
                new_pct = round(holdings[ticker]["weight"] * 100, 2)
                new_line = line.replace(weight_str, f"{new_pct}%")
                new_lines.append(new_line)
            else:
                new_lines.append(line)
        else:
            new_lines.append(line)

    return "\n".join(new_lines), ""


# ========== Main ==========

def main():
    parser = argparse.ArgumentParser(description="CIF 组合持仓自动同步")
    parser.add_argument("--dry-run", action="store_true", help="预览变更，不修改文件")
    parser.add_argument("--auto", action="store_true", help="实际执行更新")
    parser.add_argument("--from-xlsx", help="从本地 XLSX 文件读取（跳过邮件下载）")
    args = parser.parse_args()

    if not args.dry_run and not args.auto and not args.from_xlsx:
        print("请指定 --dry-run（预览）或 --auto（执行更新）")
        sys.exit(1)

    # Step 0: 加载 ticker 映射
    ticker_map = load_ticker_map()
    print(f"已加载 {len(ticker_map)} 条 ticker 映射")

    # Step 1: 获取 XLSX
    email_subject = "Portfolio H (来自邮件)"
    email_date = datetime.now().strftime("%Y-%m-%d")
    xlsx_path = args.from_xlsx

    if not xlsx_path:
        token, _, _ = get_valid_token()
        if not token:
            print("错误: 无法获取 Outlook token，请先完成授权")
            sys.exit(1)
        result = fetch_latest_xlsx(token)
        if isinstance(result, tuple) and result[0] is not None:
            xlsx_path, email_msg = result
            email_subject = email_msg.get("subject", email_subject)
            email_date = email_msg.get("receivedDateTime", email_date)[:10]
        elif isinstance(result, tuple):
            print(f"错误: {result[1]}")
            sys.exit(1)
        else:
            print(f"错误: {result}")
            sys.exit(1)

    # Step 2: 解析 HFHC
    print(f"解析 HFHC sheet: {xlsx_path}")
    hfhc_holdings, err = parse_hfhc_sheet(xlsx_path)
    if err:
        print(f"错误: {err}")
        sys.exit(1)
    print(f"  HFHC 持仓数: {len(hfhc_holdings)}")

    # Step 3: 归一化 ticker
    new_holdings, unmapped = normalize_holdings(hfhc_holdings, ticker_map)
    if unmapped:
        print(f"  未映射: {len(unmapped)} 只")

    # Step 4: 读取当前 CIF.md
    cif_holdings, cif_text = parse_cif_md()
    if not cif_holdings:
        print(f"错误: 无法解析 CIF.md")
        sys.exit(1)
    print(f"  当前 CIF.md 持仓: {len(cif_holdings)} 只")

    # Step 5: 计算变更
    changes = compute_diff(cif_holdings, new_holdings, unmapped, xlsx_path, email_subject, email_date)

    # Step 6: 输出报告
    print_diff_report(changes, unmapped, email_subject, email_date)

    # Step 7: 保存对比报告
    comparison_md = generate_comparison_md(changes, email_subject, email_date)
    reports_dir = Path.home() / "0.Agent 投研总监" / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    comparison_path = reports_dir / f"CIF对比_{datetime.now().strftime('%Y%m%d')}.md"
    with open(comparison_path, 'w', encoding='utf-8') as f:
        f.write(comparison_md)
    print(f"\n✅ 已生成: {comparison_path}")

    # Step 8: 写回文件（非 dry-run）
    if args.auto:
        removed_tickers = {c["ticker"] for c in changes if c["type"] == "removed"}
        print("更新 CIF.md...")
        new_cif_text = update_cif_md(cif_text, new_holdings, removed_tickers)
        CIF_MD_PATH.write_text(new_cif_text, encoding="utf-8")
        print(f"  ✓ {CIF_MD_PATH}")

        print("更新 CIF_组合视图.md...")
        new_view_text, err = update_cif_view_md(email_date, new_holdings)
        if new_view_text:
            CIF_VIEW_PATH.write_text(new_view_text, encoding="utf-8")
            print(f"  ✓ {CIF_VIEW_PATH}")
        else:
            print(f"  ✗ {err}")

        print(f"\n同步完成！({email_date})")

    # 清理临时文件
    if not args.from_xlsx and xlsx_path and os.path.exists(xlsx_path):
        os.unlink(xlsx_path)


if __name__ == "__main__":
    main()
