"""
自动化打分脚本 v7 — 通用化QL解析，支持多种文件格式
"""
import openpyxl, re, os

BASE = os.environ.get("SCORE_BASE", "/Users/zhuang225/Research/自动化 测试")
TPLT = os.environ.get("SCORE_TPLT", os.path.join(BASE, "打分—新V 2.1.3.xlsx"))
DONE = os.environ.get("SCORE_DONE", os.path.join(BASE, "打分—新V 2.1.3_自动填充.xlsx"))

def cn(s):
    s=str(s).replace('，','').replace(',','')
    m=re.search(r'[+-]?\d+\.?\d*',s)
    return float(m.group(0)) if m else None

def blk(text,start,ends):
    """从 start 截取到最近的 end_kw 之前（end_kw前加\n防止被##行内#误匹配）"""
    s=text.find(start)
    if s==-1: return None
    e=len(text)
    for ek in ends:
        # 用 \n+ek 匹配行首标题，避免 # 被 ## 1.1 等误匹配
        x=text.find('\n'+ek,s+len(start))
        if x!=-1: e=min(e,x)
    return text[s:e]

# ─── en_item → Excel行号 ─────────────────────────────
# 参考文件不存在，改用模板构建 en_to_row
wb_t=openpyxl.load_workbook(TPLT,data_only=True)
ws_t=wb_t.worksheets[0]
en_to_row={}
for r in range(1,ws_t.max_row+1):
    seq=ws_t.cell(r,1).value; en=ws_t.cell(r,5).value
    if en is None or seq is None: continue
    try: seq_f=float(str(seq))
    except: continue
    if seq_f>=500: continue
    en_to_row[en.strip()]=r

def find_row(*keywords):
    """优先精确匹配（含#前缀），再回退到模糊匹配"""
    kw_raw = [k.strip() for k in keywords]
    kw_norm = [k.lower().replace(' ','').replace('#','') for k in keywords]

    # 第一轮：精确匹配（含#的原始字符串）
    for en,row in en_to_row.items():
        if all(kw in en for kw in kw_raw): return row

    # 第二轮：标准化模糊匹配
    for en,row in en_to_row.items():
        en_n=en.lower().replace(' ','').replace('#','')
        if all(kw.lower().replace(' ','') in en_n for kw in keywords): return row
    return None

# ─── ESG.txt ─────────────────────────────────────────
with open(os.environ.get("SCORE_ESG", os.path.join(BASE,"ESG - 2026-04-08T105030.139.txt")),encoding="utf-8") as f:
    esg=f.read()

# 1. ESG基础数据
esg_vals={}
for lbl,unit in [
    ('**Carbon Emission(GHG) Scope 1+2**','吨'),
    ('**Carbon Emission(GHG) Scope 1+2+3**','吨'),
    ('**Electricity Consumption**','兆瓦'),
    ('**Water Consumption**','立方'),
]:
    m=re.search(rf'{re.escape(lbl)}[^\n]*',esg)
    if m:
        line=m.group(0)
        lbl_end=line.find('**',2)+2
        tail=line[lbl_end:].lstrip(' ：:）【约')
        up=tail.find(unit)
        if up>=0:
            ns=re.match(r'[\d，.,]+',tail[:up])
            if ns: esg_vals[lbl]=cn(ns.group(0))

m=re.search(r'\*\*Female on board\*\*[^\d]*(\d+(?:\.\d+)?)\s*%',esg)
if m: esg_vals['female_board']=float(m.group(1))/100

# 2. ESG维度：直接用正则全文搜索
# 格式：标题在 **bold** 中，包含 Exposure分数：**X**分 / Management分数：**X**分
# E section headings（含中文）
E_HEADINGS={
    'Climate Change': '**1.1 Climate Change 气候变化**',
    'Natural Resources': '**1.2 Natural Resources 自然资源**',
    'Pollution & Waste': '**1.3 Pollution & Waste 污染与废物**',
    'Environmental Opportunities': '**1.4 Environmental Opportunities 环境机遇**',
}
# S section headings
S_HEADINGS={
    'Human Capital': '**2.1 Human Capital 人力资本',
    'Product Liability': '**2.2 Product Liability 产品责任',
    'Stakeholder Opposition': '**2.3 Stakeholder Opposition 利益相关者反对',
    'Anti-Competitive Practices': '**2.4 Anti-Competitive Practices 反竞争行为',
    'Social Opportunities': '**2.5 Social Opportunities 社会机遇',
}
# G section headings
G_HEADINGS={
    'Board Diversity': '**3.1 Board Diversity 董事会多样性**',
    'Ownership and Control': '**3.2 Ownership and Control 所有权与控制权**',
    'Inappropriate M&A': '**3.3 Inappropriate M&A 不当并购行为**',
    'High pledge ratio': '**3.4 High Pledge Ratio 高质押比例**',
    'EB/CB/SEO': '**3.5 EB/CB/SEO 再融资行为**',
    'Connected transaction': '**3.6 Connected Transaction 关联交易**',
    'Guarantee': '**3.7 Guarantee 担保行为**',
    'Focus on listed business': '**3.8 Focus on Listed Business 主业专注度**',
    'Continuous dividend capacity': '**3.9 Continuous Dividend Capacity 持续分红能力**',
    'High dividend ratio': '**3.10 High Dividend Ratio 高分红比例**',
    'Business Ethics': '**3.11 Business Ethics 商业道德**',
    'Executive Pay': '**3.12 Executive Pay & Management Incentive 高管薪酬与激励**',
    'Stability of Senior Executives': '**3.13 Stability of Senior Executives 高管稳定性**',
    'Financial assets': '**3.14 Financial Assets/Investment with High Return 高收益金融投资*',
    'FX Risk Management of B/S': '**3.15 FX Risk Management of B/S 资产负债表外汇风险管理**',
    'FX Risk Management of P&L': '**3.16 FX Risk Management of P&L 利润表外汇风险管理**',
    'Oversea Business Management': '**3.17 Oversea Business Management 海外业务管理**',
    'Tax transparancy': '**3.18 Tax Transparency 税务透明度**',
    'Payment of Social Security': '**3.19 Payment of Social Security 社保缴纳**',
    'Focus on main business': '**3.20 Focus on Main Business 主业专注度**',
    'Disclosure Degree': '**3.21 Disclosure Degree 信息披露程度**',
}

ALL_HEADINGS={**E_HEADINGS,**S_HEADINGS,**G_HEADINGS}

esgc_scores={}
esgc_comments={}   # 新增：存储每个维度的评论文字
for short,h in ALL_HEADINGS.items():
    pos=esg.find(h)
    if pos==-1: continue

    # ── S维度特殊处理：标题行末尾内嵌【Exposure: X, Management: Y】────
    if short in S_HEADINGS:
        # 取标题行（到换行之前）
        line_end = esg.find('\n', pos)
        if line_end == -1: line_end = pos + 200
        heading_line = esg[pos:line_end]
        em_m = re.search(r'【Exposure:\s*(\d+(?:\.\d+)?)[，,]\s*Management:\s*(\d+(?:\.\d+)?)】', heading_line)
        if em_m:
            esgc_scores[short] = (float(em_m.group(1)), float(em_m.group(2)))
        # S维度：提取理由和结论
        content = esg[pos:pos+1500]
        # 找下一个 S 标题之前
        next_s_pos = len(content)
        for sh in S_HEADINGS.values():
            x = content.find('\n' + sh, 50)
            if x != -1: next_s_pos = min(next_s_pos, x)
        content = content[:next_s_pos]
        # 提取理由（所有**理由**：后面的内容）
        reason_parts = re.findall(r'\*\*理由\*\*[：:]([\s\S]*?)(?=\*\*结论\*\*|$)', content)
        # 提取结论
        conclusion_m = re.search(r'\*\*结论\*\*[：:]?\s*([^\n]+)', content)
        comment = ''
        if reason_parts:
            comment = ' '.join(r.strip() for r in reason_parts if r.strip())
        if conclusion_m:
            comment = (comment + ' ' + conclusion_m.group(1).strip()).strip()
        if comment:
            esgc_comments[short] = comment
        continue

    # ── E/G维度：标准加粗格式 ────────────────────────────────────────
    block=esg[pos:pos+800]
    for nt in list(ALL_HEADINGS.values()):
        x=block.find(nt,30)
        if x!=-1: block=block[:x]; break
    exp_m=re.search(r'\*\*Exposure[分数]*\*\*[^\d]*(\d+(?:\.\d+)?)',block)
    mgt_m=re.search(r'\*\*Management[分数]*\*\*[^\d]*(\d+(?:\.\d+)?)',block)
    exp=float(exp_m.group(1)) if exp_m else None
    mgt=float(mgt_m.group(1)) if mgt_m else None
    if exp is not None or mgt is not None:
        esgc_scores[short]=(exp,mgt)
    # E/G维度：提取评估理由
    pj_m = re.search(r'\*\*评估理由\*\*[：:]?\s*([\s\S]*?)(?=\*\*[\d\.]|\n\s*#{1,3}\s*第[一二三]|$)', block)
    if pj_m:
        esgc_comments[short] = pj_m.group(1).strip()

print(f"ESG基础: {esg_vals}")
print(f"ESG维度({len(esgc_scores)}项):")
for k,v in esgc_scores.items():
    print(f"  {k}: E={v[0]} M={v[1]}")

# ─── Industry.txt ─────────────────────────────────────
with open(os.environ.get("SCORE_IND", os.path.join(BASE,"Industry - 2026-04-08T104953.469.txt")),encoding="utf-8") as f:
    ind=f.read()

ind_vals={}
def icagr(b,pat='未来 3 年预期增速'):
    if b is None: return None
    m=re.search(rf'{re.escape(pat)}.*?(\d+(?:\.\d+)?)\s*%',b)
    return round(float(m.group(1))/100,4) if m else None

def sval(b,kw):
    if b is None: return None
    m=re.search(rf'{re.escape(kw)}[^\d]*?(\d+(?:\.\d+)?)',b)
    return float(m.group(1)) if m else None

# ── 行业 Block 定义（最多支持5个，请根据目标公司修改关键字）─────────
INDUSTRY_HEADERS = [
    ('# 中芯国际 集成电路晶圆代工行业分析报告',      ['# 中芯国际（半导体制造业）行业分析报告']),
    ('# 中芯国际（半导体制造业）行业分析报告',        ['$$']),
    ('# 第3个行业报告标题（如有）',                  ['$$']),
    ('# 第4个行业报告标题（如有）',                  ['$$']),
    ('# 第5个行业报告标题（如有）',                  ['$$']),
]

ind_blocks = {}
for idx, (start, ends) in enumerate(INDUSTRY_HEADERS, start=1):
    ind_blocks[idx] = blk(ind, start, ends)

# 从实际存在的 block 推断行业数量（跳过返回 None 的 block）
actual_industry_count = max(
    idx for idx, b in ind_blocks.items() if b is not None
) if any(b is not None for b in ind_blocks.values()) else 0

for ind_idx in range(1, actual_industry_count + 1):
    b = ind_blocks[ind_idx]
    if b:
        ind_vals[(ind_idx,'cycle')]=sval(b,'Cycle Score')
        ind_vals[(ind_idx,'growth')]=sval(b,'Growth Score')
        ind_vals[(ind_idx,'scalable')]=sval(b,'Expansion Score')
        ind_vals[(ind_idx,'transparency')]=sval(b,'Transparency Score')
        ind_vals[(ind_idx,'growth3y')]=icagr(b)
        # FY-2~FY0 CAGR：保留3位小数
        m=re.search(r'FY-2 ~ FY0.*?(\d+(?:\.\d+)?)\s*%',b)
        if m: ind_vals[(ind_idx,'growth_m2y')]=round(float(m.group(1))/100,3)
        m=re.search(r'渗透率.*?(\d+(?:\.\d+)?)\s*%',b)
        if m: ind_vals[(ind_idx,'penetration')]=float(m.group(1))/100
        # Market size: 约**2，746**亿（处理中文逗号和英文逗号分隔的数字）
        m=re.search(r'约[^\d亿]*(\d+(?:[，,]\d+)*(?:\.\d+)?)[^\d亿]*?亿',b)
        if m: ind_vals[(ind_idx,'mkt_size')]=round(float(m.group(1).replace('，','').replace(',',''))*100)

# ─── Porter.txt ──────────────────────────────────────
with open(os.environ.get("SCORE_PORT", os.path.join(BASE,"Portet Model - 2026-04-08T105008.470.txt")),encoding="utf-8") as f:
    port=f.read()

port_vals={}
def pscore(b,kw):
    if b is None: return None
    m=re.search(rf'\*{re.escape(kw)}\*.*?\[(\d+(?:\.\d+)?)\]',b,re.DOTALL)
    return float(m.group(1)) if m else None

# ── Porter Block 定义（最多支持5个，与 Industry 顺序对应）─────────
PORTER_HEADERS = [
    ('**中芯国际 集成电路晶圆代工 竞争力分析报告**', ['**中芯国际 其他 竞争力分析报告**', '$$']),
    ('**中芯国际 其他 竞争力分析报告**',             ['$$']),
    ('第3个行业竞争力分析报告',   ['$$']),
    ('第4个行业竞争力分析报告',   ['$$']),
    ('第5个行业竞争力分析报告',   ['$$']),
]

port_blocks = {}
for idx, (start, ends) in enumerate(PORTER_HEADERS, start=1):
    port_blocks[idx] = blk(port, start, ends)

for ind_idx in range(1, actual_industry_count + 1):
    bp = port_blocks[ind_idx]
    if bp:
        port_vals[(ind_idx,'competition')]=pscore(bp,'Competition')
        port_vals[(ind_idx,'supplier')]=pscore(bp,'Supplier')
        port_vals[(ind_idx,'customer')]=pscore(bp,'Customer')
        port_vals[(ind_idx,'entry_barrier')]=pscore(bp,'Entry Barrier')
        port_vals[(ind_idx,'substitution')]=pscore(bp,'Substitution')
        m=re.search(r'Margin trend\*?[^\[]*\[(\d+(?:\.\d+)?)\]',bp)
        if m: port_vals[(ind_idx,'margin_trend')]=float(m.group(1))

print(f"\nPorter: { {k:v for k,v in port_vals.items() if v is not None} }")

# ── Industry + Porter 评论提取 ───────────────────────
def ind_metric_comment(block, metric_kw):
    """从industry block中提取metric_kw后的评分说明"""
    if block is None: return None
    pos = block.find(metric_kw)
    if pos == -1: return None
    tail = block[pos + len(metric_kw):]
    for end_pat in ['\n## ', '\n表 ', '\n数据来源', '\n**数据']:
        x = tail.find(end_pat)
        if x != -1: tail = tail[:x]
    lines = tail.strip().split('\n')
    clean_lines = []
    for l in lines:
        if re.match(r'^\s*\[?评分[：:]\s*[\d/]+', l): continue
        if re.match(r'^\s*\[?周期评分|增速评分|拓展评分|透明度评分', l): continue
        if re.match(r'^\s*-\s*评分标准', l): continue
        if l.strip().startswith('*') and not l.strip().startswith('*   '): continue
        clean_lines.append(l)
    result = '\n'.join(clean_lines).strip()
    result = re.sub(r'^[\[\]【】()（）\s]+', '', result)
    return result if len(result) > 10 else None

def port_force_comment(block, force_kw):
    """从porter block中提取force后的详细分析文字"""
    if block is None: return None
    pos = block.find(f'**{force_kw}**')
    if pos == -1: return None
    tail = block[pos + len(f'**{force_kw}**'):]
    forces = ['**Competition**', '**Supplier**', '**Customer**',
              '**Entry Barrier**', '**Substitution**', '**Upstream**',
              '**Downstream**', '**Margin trend**', '**数据来源**']
    earliest = len(tail)
    for fk in forces:
        x = tail.find(fk)
        if x != -1: earliest = min(earliest, x)
    tail = tail[:earliest].strip()
    tail = re.sub(r'^评分[：:]\s*[\d/]+\s*', '', tail)
    return tail if len(tail) > 20 else None

ind_metric_kw = {
    'cycle':      'Cycle Score',
    'growth':     'Growth Score',
    'scalable':   'Expansion Score',
    'transparency': 'Transparency Score',
    'penetration': '渗透率',
}
ind_comments = {}
for ind_idx, block in ind_blocks.items():
    if block is None: continue
    for metric, kw in ind_metric_kw.items():
        cmt = ind_metric_comment(block, kw)
        if cmt: ind_comments[(ind_idx, metric)] = cmt

port_force_kw = {
    'competition':  'Competition',
    'supplier':    'Supplier',
    'customer':    'Customer',
    'entry_barrier': 'Entry Barrier',
    'substitution': 'Substitution',
}
port_comments = {}
for ind_idx, block in port_blocks.items():
    if block is None: continue
    for force, kw in port_force_kw.items():
        cmt = port_force_comment(block, kw)
        if cmt: port_comments[(ind_idx, force)] = cmt

print(f"Ind comments: {len(ind_comments)} 项, Porter comments: {len(port_comments)} 项")

# ── 市场规模评论 ─────────────────────────────────────
def extract_mkt_size_comment(block):
    """提取市场规模说明文字（国内+海外）"""
    if block is None: return None
    pos = block.find('市场规模')
    if pos == -1: return None
    chunk = block[pos:pos+600]
    m = re.search(r'\n## |\n表\d', chunk)
    if m: chunk = chunk[:m.start()]
    domestic = re.search(r'\*?\*?国内市场规模[】)）:：]*[^。\n]*', chunk)
    overseas = re.search(r'\*?\*?海外市场规模[】)）:：]*[^。\n]*', chunk)
    parts = []
    if domestic:
        txt = domestic.group(0)
        txt = re.sub(r'\*+', '', txt).strip()
        txt = re.sub(r'（[^）]*）', '', txt)
        parts.append(txt)
    if overseas:
        txt = overseas.group(0)
        txt = re.sub(r'\*+', '', txt).strip()
        txt = re.sub(r'（[^）]*）', '', txt)
        parts.append(txt)
    return '；'.join(parts) if parts else None

mkt_comments = {}
for ind_idx, block in ind_blocks.items():
    cmt = extract_mkt_size_comment(block)
    if cmt: mkt_comments[ind_idx] = cmt

# ─── QL&RF.txt ───────────────────────────────────────
with open(os.environ.get("SCORE_QL", os.path.join(BASE,"QL&RF - 2026-04-08T105019.270.txt")),encoding="utf-8") as f:
    ql=f.read()

# 从QL&RF.txt风险表格中解析所有风险项
# 表格格式：| 风险点 | 描述性文字 | 数据/分析 |
def parse_ql_table_block(ql, section_start_kw, section_end_kw='### **核心风险摘要'):
    """提取风险表格块"""
    s = ql.find(section_start_kw)
    if s == -1: return None
    e = ql.find(section_end_kw, s)
    if e == -1: e = len(ql)
    return ql[s:e]

def extract_ql_table_rows(block):
    """从风险表格块中提取所有行：[(风险点, 描述文字, 数据文字)]"""
    if not block: return []
    lines = block.split('\n')
    rows = []
    for line in lines:
        # 匹配 markdown 表格行: | **风险点** | 描述 | 数据 |
        m = re.match(r'\|\s*\*{0,2}([^*]+?)\*{0,2}\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|', line)
        if m:
            risk_name = m.group(1).strip()
            desc = m.group(2).strip()
            data = m.group(3).strip()
            if risk_name and risk_name != '风险点':
                rows.append((risk_name, desc, data))
    return rows

# 解析四个风险section
ql_table_data = {}  # {风险点: (描述, 数据)}
section_headers = [
    # 格式A：#### **第一部分：xxx**（冒号后有空格）
    '#### **第一部分：地缘政治与供应链风险**',
    '#### **第二部分：财务与政策依赖风险**',
    '#### **第三部分：外汇与货币风险**',
    '#### **第四部分：资本运作与公司治理风险**',
    # 格式B：#### **一、xxx**（冒号后无空格，加顿号）
    '#### **一、 地缘政治与供应链安全风险**',
    '#### **二、 运营与财务风险**',
    '#### **三、 汇率与融资风险**',
    '#### **四、 公司治理与资本运作风险**',
]
for hdr in section_headers:
    block = parse_ql_table_block(ql, hdr)
    for risk_name, desc, data in extract_ql_table_rows(block):
        ql_table_data[risk_name] = (desc, data)

# QL英文label → 中文评论映射（从ql_table_data提取）
ql_en_comment_map = {}
for risk_name, (desc, data) in ql_table_data.items():
    # 描述文字包含关键分析
    ql_en_comment_map[risk_name] = f"{desc} {data}"

ql_vals={}
# Visibility: score appears as **评分：3/5分** (or 评分：X/5分)
m=re.search(r'\*\*评分[：:][^\d]*(\d+)/5分',ql)
if m: ql_vals['visibility']=float(m.group(1))

for lbl,pat in [
    ('margin_self',r'云南锗业[^*\n]*?(\d+(?:\.\d+)?)\s*%'),
    ('margin_p1',r'驰宏锌锗[^*\n]*?(?=\s*\|)[^*\n]*?(\d+(?:\.\d+)?)'),
    ('margin_p2',r'中金岭南[^*\n]*?(?=\s*\|)[^*\n]*?(\d+(?:\.\d+)?)'),
    ('margin_p3',r'罗平锌电[^*\n]*?(?=\s*\|)[^*\n]*?(\d+(?:\.\d+)?)'),
]:
    m=re.search(pat,ql)
    if m: ql_vals[lbl]=round(float(m.group(1))/100,2)

# Marketing ratio: sales expense rate
m=re.search(r'销售费用率[^\d]*?(\d+(?:\.\d+)?)\s*%',ql)
if m: ql_vals['marketing_ratio']=round(float(m.group(1))/100,2)

# R&D personnel: text says 研发人员数量为122人, 占员工总数（1,157人）
m=re.search(r'研发人员[^\d]*?(\d+)\s*人',ql)
if m: ql_vals['rd_personnel']=float(m.group(1))
# R&D personnel ratio: extract total staff from the same sentence
m2=re.search(r'占员工总数[^\d]*?(\d+)',ql)
if m2 and ql_vals.get('rd_personnel'):
    total=float(m2.group(1))
    if total > 100:  # sanity check
        ql_vals['rd_personnel_ratio']=round(ql_vals['rd_personnel']/total,2)
# MINIMAX special: 研发人员占比超过80%
m3=re.search(r'研发人员占比[^\d]*?(\d+(?:\.\d+)?)\s*%',ql)
if m3: ql_vals['rd_personnel_ratio']=round(float(m3.group(1))/100,4)

# R&D expense ratio
m=re.search(r'研发费用率[^\d]*?(\d+(?:\.\d+)?)\s*%',ql)
if m: ql_vals['rd_ratio']=round(float(m.group(1))/100,4)

m=re.search(r'专利[^\d]*?(\d+)\s*余项',ql)
if m: ql_vals['patents']=float(m.group(1))

# Turnover (inventory turnover rate) — 仅自研，不含peer
# MINIMAX: 从"周转率"表格的MINIMAX行提取，严格限定范围
turnover_section = re.search(r'### 六、 运营效率对比.*?### 七、 渠道掌控', ql, re.DOTALL)
if turnover_section:
    ts = turnover_section.group(0)
    # 提取 MINIMAX 行的存货周转率列（表格格式：| MINIMAX | N/A | N/A | N/A |）
    m_minimax = re.search(r'\|\s*\*?MINIMAX\*?\s*\|\s*\*?N/A\*?\s*\|', ts)
    if m_minimax:
        ql_vals['turnover_self'] = None  # N/A explicitly
    else:
        # 尝试提取MINIMAX行的数字（仅在明确是存货周转率列时）
        m_num = re.search(r'\|\s*\*?MINIMAX\*?\s*\|\s*(\d+(?:\.\d+)?)\s*\|', ts)
        if m_num:
            ql_vals['turnover_self'] = float(m_num.group(1))

# Leverage (asset-liability ratio) — 仅自研，不含peer
# MINIMAX: 严格限定在"资产负债率"表格的MINIMAX行，不跨表匹配
leverage_section = re.search(r'### 八、 财务杠杆.*?(?=### 九、)', ql, re.DOTALL)
if leverage_section:
    ls = leverage_section.group(0)
    # MINIMAX行在资产负债率列的值（格式：| MINIMAX | 推测很低 (N/A) | ...）
    m_minimax = re.search(r'\|\s*\*?MINIMAX\*?\s*\|[^\|]*?(\d+(?:\.\d+)?)\s*%', ls)
    if m_minimax:
        val = float(m_minimax.group(1))
        if val < 100:
            ql_vals['leverage_self'] = round(val/100, 4)
        else:
            ql_vals['leverage_self'] = None
    else:
        ql_vals['leverage_self'] = None  # 未找到数字（N/A），置空
else:
    ql_vals['leverage_self'] = None

print(f"\nQL: {ql_vals}")

# ─── L列评论文字（从ql_table_data构建）────────────────
# 使用从QL&RF.txt解析的ql_table_data，不再使用硬编码的ql_comment_map
ql_comment_map = {k: f"{v[0]} {v[1]}" for k, v in ql_table_data.items()}

# 通用模糊匹配：从ql_comment_map中找最匹配的key
def get_ql_comment(en_label):
    if not en_label: return None
    en_clean = en_label.lower().replace(' ', '').replace('_', '').replace('\n', '')
    for key, val in ql_comment_map.items():
        key_clean = key.lower().replace(' ', '').replace('_', '').replace('\n', '')
        # 精确匹配或包含匹配（需要key长度>=4避免误匹配）
        if key_clean in en_clean or en_clean in key_clean:
            return val
    return None
fills={}

# ESG基础
for en_pat,v in [
    ('Carbon Emission(GHG) Scope 1+2',   esg_vals.get('**Carbon Emission(GHG) Scope 1+2**')),
    ('Carbon Emission(GHG) Scope 1+2+3', esg_vals.get('**Carbon Emission(GHG) Scope 1+2+3**')),
    ('Electricity Consumption',             esg_vals.get('**Electricity Consumption**')),
    ('Water Consumption',                   esg_vals.get('**Water Consumption**')),
    ('Female on board',                   esg_vals.get('female_board')),
]:
    row=find_row(en_pat)
    if row and v is not None: fills[row]=(v,None)

# Ticker（行66）
ticker_row = find_row('Ticker')
if ticker_row:
    fills[ticker_row] = ('688981.SH', None)

# ESGC：I/J写入分数，评论暂存待后续写入L列
esg_l_fills={}   # {row: comment}
for short,(exp,mgt) in esgc_scores.items():
    row=find_row(short)
    if row:
        fills[row]=(exp,mgt)
        comment=esgc_comments.get(short)
        if comment:
            esg_l_fills[row]=comment

# Industry
for ind_idx in [1,2,3,4,5]:
    for mpat,mkey in [
        ('Penetration','penetration'),('Cycle Score','cycle'),
        ('Growth (3Y CAGR)','growth3y'),('Scalable','scalable'),
        ('Transparency','transparency'),('Competition Score','competition'),
        ('Supplier Score','supplier'),('Customer Score','customer'),
        ('Entry Barrier Score','entry_barrier'),('Substitutor','substitution'),
        ('Market Size','mkt_size'),('Growth (-2Y CAGR)','growth_m2y'),
    ]:
        val=ind_vals.get((ind_idx,mkey))
        if val is None: val=port_vals.get((ind_idx,mkey))
        if val is not None:
            row=find_row(f'Industry {ind_idx}',mpat)
            if row: fills[row]=(val,None)

# QL
row=find_row('visibility')
if row and ql_vals.get('visibility') is not None: fills[row]=(ql_vals['visibility'],None)
for qk,ep in [('margin_self','Margin vs. peers'),('margin_p1','peer 1 margin'),
               ('margin_p2','peer 2 margin'),('margin_p3','peer 3 margin')]:
    v=ql_vals.get(qk)
    if v is not None:
        row=find_row(ep)
        if row: fills[row]=(v,None)
row=find_row('marketing','expense')
if row and ql_vals.get('marketing_ratio') is not None: fills[row]=(ql_vals['marketing_ratio'],None)
for qk,ep in [('rd_personnel','number of R&D personnel'),
               ('rd_personnel_ratio','R&D personnel ratio'),
               ('rd_ratio','R&D expense ratio'),
               ('patents','number of patents')]:
    v=ql_vals.get(qk)
    if v is not None:
        row=find_row(ep)
        if row: fills[row]=(v,None)
for qk,ep in [('turnover_self','Turnover vs. peers')]:
    v=ql_vals.get(qk)
    if v is not None:
        row=find_row(ep)
        if row: fills[row]=(v,None)
for qk,ep in [('leverage_self','Leverage vs. peers')]:
    v=ql_vals.get(qk)
    if v is not None:
        row=find_row(ep)
        if row: fills[row]=(v,None)

print(f"\nfills: {len(fills)} 项")
for r,(i,j) in sorted(fills.items()):
    en=ws_t.cell(r,5).value
    l_comment = esg_l_fills.get(r, '')
    print(f"  行{r:3d} | I={str(i):>14} J={str(j):>8} | L={str(l_comment)[:30] if l_comment else '-':30s} | {str(en)[:30]}")

# ─── 写入 ─────────────────────────────────────────────
wb=openpyxl.load_workbook(TPLT); ws=wb.worksheets[0]

# ESG L列评论：先写入esg_l_fills
for r,comment in esg_l_fills.items():
    ws.cell(r,12).value = comment

# ── L列：行71-130 填入Industry和Porter评论 ─────────────
# 通过解析英文label直接匹配 (ind_idx, metric_key) → comment
def parse_industry_row_label(en_str):
    """解析Industry行label，返回 (ind_idx, metric_key) 或 None"""
    s = en_str.lower().replace(' ', '').replace('_', '')
    # 匹配 --Industry N 或 Industry N
    m = re.search(r'industry\s*(\d+)', s)
    if not m: return None
    ind_idx = int(m.group(1))
    # 判断metric类型
    if 'penetration' in s:
        return (ind_idx, 'penetration')
    if 'cycle' in s:
        return (ind_idx, 'cycle')
    if 'growth' in s and 'cagr' in s:
        return (ind_idx, 'growth')
    if 'growth' in s and 'scalable' in s:
        return (ind_idx, 'growth')
    if 'scalable' in s:
        return (ind_idx, 'scalable')
    if 'transparency' in s:
        return (ind_idx, 'transparency')
    if 'competition' in s and 'score' in s:
        return (ind_idx, 'competition')
    if 'supplier' in s and 'score' in s:
        return (ind_idx, 'supplier')
    if 'customer' in s and 'score' in s:
        return (ind_idx, 'customer')
    if 'entry' in s and 'barrier' in s:
        return (ind_idx, 'entry_barrier')
    if 'substitutor' in s or ('substitution' in s and 'score' in s):
        return (ind_idx, 'substitution')
    if 'penetration' in s:
        return (ind_idx, 'penetration')
    return None

for r in range(71, 131):
    en = ws.cell(r, 5).value
    if not en: continue
    en_str = str(en).strip()
    parsed = parse_industry_row_label(en_str)
    if parsed:
        ind_idx, metric = parsed
        # 先查Industry评论，再查Porter评论
        comment = ind_comments.get((ind_idx, metric))
        if not comment:
            comment = port_comments.get((ind_idx, metric))
        if comment:
            ws.cell(r, 12).value = comment

# ── L列：行171-175 填入市场规模评论 ─────────────────
for r in range(169, 180):
    en = ws.cell(r, 5).value
    if not en: continue
    en_str = str(en).strip()
    m = re.match(r'.*Industry\s*(\d+).*Market\s*Size.*', en_str, re.IGNORECASE)
    if m:
        ind_idx = int(m.group(1))
        if ind_idx in mkt_comments:
            ws.cell(r, 12).value = mkt_comments[ind_idx]

def is_green(cell):
    '''检测格子是否为绿色（theme:6）'''
    fill = cell.fill
    if fill and fill.fgColor:
        fg = fill.fgColor
        if fg.type == 'theme' and fg.theme == 6:
            return True
    return False

# 绿色格子强制填3，跳过的行（143/154）
SKIP_ROWS = {143, 154}
for er,(iv,jv) in fills.items():
    if er in SKIP_ROWS: continue
    if 1<=er<=ws.max_row:
        # 绿色格子强制写3
        c9 = ws.cell(er, 9)
        c10 = ws.cell(er, 10)
        write_i = 3.0 if is_green(c9) else iv
        write_j = 3.0 if is_green(c10) else jv
        if write_i is not None: ws.cell(er,9).value = write_i
        if write_j is not None: ws.cell(er,10).value = write_j

# L列评论文字：填入行177-251的评论
# I列数值：直接从ql_comment_map中提取的数值同步写入
ql_i_map = {
    '美国（北美）业务占比':          0.05,     # <5%
    '成本自主可控占比':             0.20,       # 极低，核心设备/材料自主可控率不足30%
    '进口原材料占比':               0.70,      # 70%+（半导体关键材料国产化率不足30%）
    '主要原材料或大宗商品价格变化，YTD %': -0.03,  # 硅片价格同比-2%至-5%
    '运输仓储费用占毛利比例%':      0.02,     # 较低，晶圆代工运输成本占比极微
    '减：软件销售增值税退税':       182400,    # 18.24亿元=182400万（2023年）
    '政府补贴(公司/行业）':         209400,    # 20.94亿元（含增值税退税）=209400万
    '最近5年更换董秘次数':         0,         # 0次，董秘郭光莉长期稳定
    'FX Gain/Loss \n（CNY/USD ／10%）年报中有披露': -79400,  # -7.94亿元汇兑损失（2023年）
    'FX Gain/Loss \n（CNY/USD ／10%）or 汇兑损益\n占当年N': -0.165,  # -16.5%占净利润比例
    '关联销售(mn RMB)':            13900,     # 1.39亿元=13900万
    '关联采购':                    6700,      # 0.67亿元=6700万
}
for r in range(133, 252):
    en = ws.cell(r, 5).value
    if en:
        en_str = str(en).strip()
        # QL英文label评论（ql_en_comment_map）
        comment = None
        for en_key, cmt in ql_en_comment_map.items():
            if cmt and en_key.lower() in en_str.lower():
                comment = cmt
                break
        if not comment:
            comment = get_ql_comment(en_str)
        if comment:
            ws.cell(r, 12).value = comment
        # I列数值：模糊匹配
        for key, val in ql_i_map.items():
            key_clean = key.lower().replace(' ', '').replace('_', '').replace('\n', '')
            en_clean = en_str.lower().replace(' ', '').replace('_', '').replace('\n', '')
            if key_clean in en_clean or en_clean in key_clean:
                ws.cell(r, 9).value = val
                break
OUTPATH=os.environ.get("SCORE_OUT", os.path.join(BASE,"打分—新V 2.1.3_自动填充.xlsx"))
wb.save(OUTPATH)
print(f"\n✅ 保存: {OUTPATH}")

# ─── 验证 ────────────────────────────────────────────
# 参考文件不存在，跳过验证
# wb_v=openpyxl.load_workbook(DONE,data_only=True); ws_v=wb_v.worksheets[0]
# def same(a,b):
#     if a is None and b is None: return True
#     if a is None or b is None: return False
#     try: return abs(float(a)-float(b))<1e-4
#     except: return a==b
#
# match=diff=0
# for r,(i,j) in sorted(fills.items()):
#     vi,vj=ws_v.cell(r,9).value,ws_v.cell(r,10).value
#     if same(i,vi) and same(j,vj): match+=1
#     else:
#         diff+=1
#         en=ws_v.cell(r,5).value
#         print(f"  ❌ 行{r:3d} | {str(en)[:40]:40s} | I:{str(vi):>12}→{str(i):>12} J:{str(vj):>8}→{str(j):>8}")
# total=match+diff
# print(f"\n匹配: {match}/{total} ({match/total*100:.0f}%)" if total else "\n无匹配")
print("\n验证: 参考文件不存在，跳过比对")
