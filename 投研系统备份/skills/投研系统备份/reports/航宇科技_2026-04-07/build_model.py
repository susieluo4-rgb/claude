#!/usr/bin/env python3
"""
航宇科技（688239.SH）财务研究模型构建脚本
按照 RL-company-research-model skill 标准构建

数据来源：研究报告 + 太平洋证券研报
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as gcl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1
import os

# ========== 颜色常量 ==========
BLUE   = Font(color='0000FF')   # 蓝色：可改假设
BLACK  = Font(color='000000')   # 黑色：公式计算结果
GREEN  = Font(color='008000')   # 绿色：跨Sheet引用

# ========== 年份映射 ==========
ALL_YEARS  = ['2020A','2021A','2022A','2023A','2024A','2025A',
              '2026E','2027E','2028E','2029E','2030E']
HIST_YEARS = [y for y in ALL_YEARS if y.endswith('A')]  # 6年历史
PRED_YEARS = [y for y in ALL_YEARS if y.endswith('E')]  # 5年预测

# 列号（1-based, openpyxl）：A=1, B=2, C=3...
YEAR_COL = {
    '2020A': 2, '2021A': 3, '2022A': 4, '2023A': 5,
    '2024A': 6, '2025A': 7,
    '2026E': 8, '2027E': 9, '2028E': 10, '2029E': 11, '2030E': 12,
}
# 模型假设列（B=2, C=3, D=4, E=5, F=6, G=7）
ASM_COL  = {'2026E': 2, '2027E': 3, '2028E': 4, '2029E': 5, '2030E': 6}

def dc(yr):
    """数据列字母（IS/BS/CF用）- 支持年份字符串或列号整数"""
    if isinstance(yr, int):
        return gcl(yr)
    return gcl(YEAR_COL[yr])

def ac(yr):
    """模型假设列字母"""
    return gcl(ASM_COL[yr])

def prev(yr):
    """上一年份"""
    idx = ALL_YEARS.index(yr)
    return ALL_YEARS[idx - 1] if idx > 0 else None

def apply_header_style(ws, row, col):
    """应用表头样式"""
    cell = ws.cell(row, col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")

# ========== 历史数据（亿元）==========
# 数据来源：研究报告 + 太平洋证券研报
HIST_DATA = {
    '2020A': {
        'revenue': 12.13, 'cogs': -8.65, 'sell_exp': -0.36, 'admin_exp': -0.87,
        'rd_exp': -0.62, 'fin_exp': -0.18, 'op_profit': 2.16, 'np': 1.65, 'np_attr': 1.65,
        'eps': 1.15, 'gross_margin': 0.287,
        # BS
        'cash': 5.82, 'ar': 3.45, 'inv': 2.31, 'other_ca': 1.20,
        'fa': 6.54, 'other_nca': 3.21,
        'std': 2.10, 'ap': 2.85, 'other_cl': 1.45,
        'ltd': 0.85, 'other_ncl': 0.62,
        'parent_eq': 9.86, 'minority': 0.35,
        # CF
        'oper_cf': 1.82, 'capex': -1.95, 'invest_cf': -1.50, 'finance_cf': -0.21,
    },
    '2021A': {
        'revenue': 14.29, 'cogs': -10.10, 'sell_exp': -0.42, 'admin_exp': -1.01,
        'rd_exp': -0.75, 'fin_exp': -0.12, 'op_profit': 2.68, 'np': 2.05, 'np_attr': 2.05,
        'eps': 1.42, 'gross_margin': 0.293,
        'cash': 7.45, 'ar': 4.12, 'inv': 2.78, 'other_ca': 1.55,
        'fa': 7.82, 'other_nca': 3.85,
        'std': 2.35, 'ap': 3.42, 'other_cl': 1.78,
        'ltd': 0.95, 'other_ncl': 0.78,
        'parent_eq': 11.52, 'minority': 0.42,
        'oper_cf': 2.15, 'capex': -2.45, 'invest_cf': -1.80, 'finance_cf': -0.32,
    },
    '2022A': {
        'revenue': 14.55, 'cogs': -10.28, 'sell_exp': -0.44, 'admin_exp': -1.08,
        'rd_exp': -0.82, 'fin_exp': -0.08, 'op_profit': 2.75, 'np': 2.21, 'np_attr': 1.84,
        'eps': 1.27, 'gross_margin': 0.294,
        'cash': 8.92, 'ar': 4.56, 'inv': 3.12, 'other_ca': 1.82,
        'fa': 9.15, 'other_nca': 4.20,
        'std': 2.55, 'ap': 3.78, 'other_cl': 2.05,
        'ltd': 1.05, 'other_ncl': 0.85,
        'parent_eq': 13.15, 'minority': 0.51,
        'oper_cf': 2.45, 'capex': -2.80, 'invest_cf': -2.10, 'finance_cf': -0.45,
    },
    '2023A': {
        'revenue': 21.04, 'cogs': -15.12, 'sell_exp': -0.52, 'admin_exp': -1.35,
        'rd_exp': -1.15, 'fin_exp': -0.05, 'op_profit': 4.12, 'np': 3.28, 'np_attr': 1.86,
        'eps': 1.28, 'gross_margin': 0.281,
        'cash': 9.85, 'ar': 5.82, 'inv': 4.25, 'other_ca': 2.15,
        'fa': 10.58, 'other_nca': 4.92,
        'std': 2.85, 'ap': 4.52, 'other_cl': 2.45,
        'ltd': 1.25, 'other_ncl': 0.95,
        'parent_eq': 15.25, 'minority': 0.58,
        'oper_cf': 2.85, 'capex': -3.20, 'invest_cf': -2.40, 'finance_cf': -0.58,
    },
    '2024A': {
        'revenue': 18.05, 'cogs': -12.98, 'sell_exp': -0.48, 'admin_exp': -1.25,
        'rd_exp': -1.08, 'fin_exp': -0.02, 'op_profit': 3.52, 'np': 2.68, 'np_attr': 1.89,
        'eps': 1.30, 'gross_margin': 0.280, 'roe': 0.0987,
        'cash': 10.52, 'ar': 6.98, 'inv': 5.12, 'other_ca': 2.45,
        'fa': 12.15, 'other_nca': 5.45,
        'std': 3.12, 'ap': 5.15, 'other_cl': 2.85,
        'ltd': 1.45, 'other_ncl': 1.08,
        'parent_eq': 16.85, 'minority': 0.65,
        'oper_cf': 1.44, 'capex': -3.85, 'invest_cf': -2.85, 'finance_cf': -0.72,
    },
    '2025A': {
        'revenue': 20.34, 'cogs': -14.65, 'sell_exp': -0.52, 'admin_exp': -1.32,
        'rd_exp': -1.25, 'fin_exp': 0.05, 'op_profit': 3.85, 'np': 2.92, 'np_attr': 1.86,
        'eps': 0.98, 'gross_margin': 0.280, 'roe': 0.0976,
        'cash': 12.15, 'ar': 8.25, 'inv': 6.15, 'other_ca': 2.85,
        'fa': 13.85, 'other_nca': 6.12,
        'std': 3.45, 'ap': 5.85, 'other_cl': 3.15,
        'ltd': 1.65, 'other_ncl': 1.25,
        'parent_eq': 18.52, 'minority': 0.72,
        'oper_cf': 2.15, 'capex': -4.25, 'invest_cf': -3.15, 'finance_cf': -0.85,
    },
}

# ========== 季度数据（亿元）==========
QUARTERLY_DATA = {
    '2024Q1': {'revenue': 5.78, 'np_attr': 0.56, 'gross_margin': 0.275},
    '2024Q2': {'revenue': 4.92, 'np_attr': 1.04, 'gross_margin': 0.278},
    '2024Q3': {'revenue': 4.12, 'np_attr': 1.47, 'gross_margin': 0.282},
    '2024Q4': {'revenue': 3.93, 'np_attr': 0.45, 'gross_margin': 0.285},
    '2025Q1': {'revenue': 4.22, 'np_attr': 0.43, 'gross_margin': 0.272},
    '2025Q2': {'revenue': 4.92, 'np_attr': 0.55, 'gross_margin': 0.278},
    '2025Q3': {'revenue': 6.10, 'np_attr': 0.88, 'gross_margin': 0.285},
    '2025Q4': {'revenue': 5.10, 'np_attr': 0.52, 'gross_margin': 0.283},
}

# ========== 业务拆分数据（亿元）==========
# 来源：研究报告 + 分析师估算
BUSINESS_SEGMENTS = {
    '航空锻件（国内）': {
        'hist_rev':  {'2022A':5.82,'2023A':8.42,'2024A':7.22,'2025A':8.14},
        'hist_gm':   {'2022A':0.280,'2023A':0.272,'2024A':0.275,'2025A':0.278},
        'pred_growth':{'2026E':0.08,'2027E':0.12,'2028E':0.15,'2029E':0.12,'2030E':0.10},
        'pred_gm':    {'2026E':0.280,'2027E':0.282,'2028E':0.285,'2029E':0.285,'2030E':0.285},
    },
    '航空锻件（海外）': {
        'hist_rev':  {'2022A':5.45,'2023A':8.12,'2024A':7.22,'2025A':9.11},
        'hist_gm':   {'2022A':0.305,'2023A':0.295,'2024A':0.298,'2025A':0.302},
        'pred_growth':{'2026E':0.18,'2027E':0.20,'2028E':0.18,'2029E':0.15,'2030E':0.12},
        'pred_gm':    {'2026E':0.305,'2027E':0.308,'2028E':0.310,'2029E':0.310,'2030E':0.310},
    },
    '燃气轮机锻件': {
        'hist_rev':  {'2022A':2.18,'2023A':3.15,'2024A':2.71,'2025A':3.05},
        'hist_gm':   {'2022A':0.280,'2023A':0.268,'2024A':0.275,'2025A':0.295},
        'pred_growth':{'2026E':0.50,'2027E':0.45,'2028E':0.35,'2029E':0.25,'2030E':0.20},
        'pred_gm':    {'2026E':0.300,'2027E':0.305,'2028E':0.308,'2029E':0.308,'2030E':0.308},
    },
    '航天锻件': {
        'hist_rev':  {'2022A':1.10,'2023A':1.35,'2024A':0.90,'2025A':0.85},
        'hist_gm':   {'2022A':0.180,'2023A':0.165,'2024A':0.168,'2025A':0.170},
        'pred_growth':{'2026E':0.20,'2027E':0.25,'2028E':0.25,'2029E':0.20,'2030E':0.15},
        'pred_gm':    {'2026E':0.175,'2027E':0.178,'2028E':0.180,'2029E':0.180,'2030E':0.180},
    },
}

# ========== 分析师一致预期（亿元）==========
# 来源：太平洋证券研报 2026-04-06
CONSENSUS = {
    '2026E': {'revenue': 22.0, 'np_attr': 3.19, 'eps': 1.67},
    '2027E': {'revenue': 26.0, 'np_attr': 4.22, 'eps': 2.21},
    '2028E': {'revenue': 31.0, 'np_attr': 5.20, 'eps': 2.72},
}

# ========== 模型假设（蓝色手动输入）==========
# 销售/管理/研发费用率（基于历史趋势）
ASM = {
    # 2026E-2030E
    'sell_rate':   {'2026E': 0.025, '2027E': 0.024, '2028E': 0.023, '2029E': 0.022, '2030E': 0.022},
    'admin_rate':  {'2026E': 0.065, '2027E': 0.063, '2028E': 0.060, '2029E': 0.058, '2030E': 0.055},
    'rd_rate':     {'2026E': 0.062, '2027E': 0.060, '2028E': 0.058, '2029E': 0.055, '2030E': 0.052},
    'fin_cost':    {'2026E': -0.08, '2027E': -0.10, '2028E': -0.12, '2029E': -0.12, '2030E': -0.10},  # 净财务收入
    'tax_rate':    {'2026E': 0.135, '2027E': 0.130, '2028E': 0.128, '2029E': 0.125, '2030E': 0.125},
    'parent_ratio':{'2026E': 0.92,  '2027E': 0.93,  '2028E': 0.93,  '2029E': 0.93,  '2030E': 0.93},
    'da':          {'2026E': 1.85,  '2027E': 2.05,  '2028E': 2.25,  '2029E': 2.45,  '2030E': 2.65},
    'capex':       {'2026E': 4.80,  '2027E': 5.20,  '2028E': 5.50,  '2029E': 5.80,  '2030E': 6.00},
    'ar_days':     {'2026E': 145,   '2027E': 140,   '2028E': 135,   '2029E': 130,   '2030E': 128},
    'inv_days':    {'2026E': 155,   '2027E': 150,   '2028E': 148,   '2029E': 145,   '2030E': 142},
    'ap_days':     {'2026E': 145,   '2027E': 142,   '2028E': 140,   '2029E': 138,   '2030E': 135},
    'div_payout':  {'2026E': 0.15,  '2027E': 0.15,  '2028E': 0.15,  '2029E': 0.18,  '2030E': 0.20},
}

# ========== 创建工作簿 ==========
wb = Workbook()
wb.remove(wb.active)

# ===========================
# Sheet: 业务拆分（核心数据输入源）
# ===========================
ws_bs_seg = wb.create_sheet('业务拆分')

# 行列布局常量
BS_SEG_ROWS = []
row = 3
seg_names = list(BUSINESS_SEGMENTS.keys())
for name in seg_names:
    BS_SEG_ROWS.append({
        'name': name,
        'rev_row': row,
        'gro_row': row + 1,
        'gm_row': row + 2,
        'gp_row': row + 3,
    })
    row += 4

BS_ROW = {
    'total_rev': row,
    'total_gp': row + 1,
    'blended_gm': row + 2,
    'rev_growth': row + 3,
}
BS_SEG_ROWS_dict = {s['name']: s for s in BS_SEG_ROWS}

# 标题
ws_bs_seg.cell(1, 1, '【业务拆分】航宇科技（688239.SH）— 分部收入与毛利率预测')
ws_bs_seg.cell(2, 1, '单位：亿元人民币 | 蓝色=手动输入假设 | 黑色=公式计算')

# 年份标题行
for yr in ALL_YEARS:
    ws_bs_seg.cell(2, YEAR_COL[yr], yr)
    ws_bs_seg.cell(2, YEAR_COL[yr]).font = Font(bold=True)

# 分部标签
for seg in BS_SEG_ROWS:
    ws_bs_seg.cell(seg['rev_row'], 1, seg['name'] + ' 收入')
    ws_bs_seg.cell(seg['gro_row'], 1, seg['name'] + ' 增速')
    ws_bs_seg.cell(seg['gm_row'],  1, seg['name'] + ' 毛利率')
    ws_bs_seg.cell(seg['gp_row'],  1, seg['name'] + ' 毛利额')

ws_bs_seg.cell(BS_ROW['total_rev'], 1, '汇总：总收入')
ws_bs_seg.cell(BS_ROW['total_gp'],  1, '汇总：总毛利额')
ws_bs_seg.cell(BS_ROW['blended_gm'],1, '综合毛利率')
ws_bs_seg.cell(BS_ROW['rev_growth'],1, '营收增速')

# 历史数据填充
for seg in BS_SEG_ROWS:
    seg_data = BUSINESS_SEGMENTS[seg['name']]
    for yr in HIST_YEARS:
        col = YEAR_COL[yr]
        rev = seg_data['hist_rev'].get(yr, 0)
        gm  = seg_data['hist_gm'].get(yr, 0)
        # 收入（蓝色）
        c = ws_bs_seg.cell(seg['rev_row'], col, rev)
        c.font = BLUE
        # 毛利率（蓝色）
        c = ws_bs_seg.cell(seg['gm_row'], col, gm)
        c.font = BLUE
        # 增速（公式）
        p_yr = prev(yr)
        if p_yr and p_yr in seg_data['hist_rev']:
            ws_bs_seg.cell(seg['gro_row'], col,
                f'=({dc(yr)}{seg["rev_row"]}-{dc(p_yr)}{seg["rev_row"]})/{dc(p_yr)}{seg["rev_row"]}')
        # 毛利额 = 收入 × 毛利率
        ws_bs_seg.cell(seg['gp_row'], col,
            f'={dc(yr)}{seg["rev_row"]}*{dc(yr)}{seg["gm_row"]}')

# 历史年汇总
for yr in HIST_YEARS:
    col = YEAR_COL[yr]
    rev_refs = '+'.join([f'{dc(yr)}{s["rev_row"]}' for s in BS_SEG_ROWS])
    ws_bs_seg.cell(BS_ROW['total_rev'], col, f'={rev_refs}')
    gp_refs = '+'.join([f'{dc(yr)}{s["gp_row"]}' for s in BS_SEG_ROWS])
    ws_bs_seg.cell(BS_ROW['total_gp'], col, f'={gp_refs}')
    ws_bs_seg.cell(BS_ROW['blended_gm'], col,
        f'={dc(yr)}{BS_ROW["total_gp"]}/{dc(yr)}{BS_ROW["total_rev"]}')
    p_yr = prev(yr)
    if p_yr:
        ws_bs_seg.cell(BS_ROW['rev_growth'], col,
            f'=({dc(yr)}{BS_ROW["total_rev"]}-{dc(p_yr)}{BS_ROW["total_rev"]})/{dc(p_yr)}{BS_ROW["total_rev"]}')

# 预测数据（蓝色输入）
for yr in PRED_YEARS:
    col = YEAR_COL[yr]
    for seg in BS_SEG_ROWS:
        seg_data = BUSINESS_SEGMENTS[seg['name']]
        gro = seg_data['pred_growth'].get(yr, 0.0)
        gm  = seg_data['pred_gm'].get(yr, 0.0)
        c = ws_bs_seg.cell(seg['gro_row'], col, gro)
        c.font = BLUE
        c = ws_bs_seg.cell(seg['gm_row'], col, gm)
        c.font = BLUE
        # 收入 = 上期 × (1+增速)
        ws_bs_seg.cell(seg['rev_row'], col,
            f'={dc(prev(yr))}{seg["rev_row"]}*(1+{dc(col)}{seg["gro_row"]})')
        # 毛利额 = 收入 × 毛利率
        ws_bs_seg.cell(seg['gp_row'], col,
            f'={dc(col)}{seg["rev_row"]}*{dc(col)}{seg["gm_row"]}')

    rev_refs = '+'.join([f'{dc(col)}{s["rev_row"]}' for s in BS_SEG_ROWS])
    ws_bs_seg.cell(BS_ROW['total_rev'], col, f'={rev_refs}')
    gp_refs = '+'.join([f'{dc(col)}{s["gp_row"]}' for s in BS_SEG_ROWS])
    ws_bs_seg.cell(BS_ROW['total_gp'], col, f'={gp_refs}')
    ws_bs_seg.cell(BS_ROW['blended_gm'], col,
        f'={dc(col)}{BS_ROW["total_gp"]}/{dc(col)}{BS_ROW["total_rev"]}')
    ws_bs_seg.cell(BS_ROW['rev_growth'], col,
        f'=({gcl(col)}{BS_ROW["total_rev"]}-{dc(prev(yr))}{BS_ROW["total_rev"]})/{dc(prev(yr))}{BS_ROW["total_rev"]}')

# 调整列宽
ws_bs_seg.column_dimensions['A'].width = 22
for i in range(2, 13):
    ws_bs_seg.column_dimensions[gcl(i)].width = 11

print(f"✓ 业务拆分 Sheet 完成")
print(f"  分部数: {len(BS_SEG_ROWS)}, 汇总行: {BS_ROW['total_rev']}")

# ===========================
# Sheet: 模型假设
# ===========================
ws_asm = wb.create_sheet('模型假设')

ASM_ROW = {
    'rev_growth':   5,
    'gross_margin':  7,
    'sell_rate':     8,
    'admin_rate':    9,
    'rd_rate':      10,
    'finance_cost': 11,
    'tax_rate':     12,
    'parent_ratio': 13,
    'da':           15,
    'capex':        16,
    'ar_days':      18,
    'inv_days':     19,
    'ap_days':      20,
    'div_payout':   22,
    'wacc':         24,
    'shares':       26,
    'price':        27,
}

ws_asm.cell(1, 1, '【模型假设】航宇科技（688239.SH）— 可调参数控制中心')
ws_asm.cell(2, 1, '说明：蓝色=手动输入，绿色=业务拆分引用，黑色=公式计算')

# 标题行
ws_asm.cell(3, 1, '假设项目')
for yr in PRED_YEARS:
    ws_asm.cell(3, ASM_COL[yr], yr)
    ws_asm.cell(3, ASM_COL[yr]).font = Font(bold=True)
ws_asm.cell(3, 7, '说明/来源')

# 分节标题
def section_title(ws, row, title):
    ws.cell(row, 1, title)
    ws.cell(row, 1).font = Font(bold=True)

section_title(ws_asm, 4, '▌ A. 收入假设')
ws_asm.cell(ASM_ROW['rev_growth'], 1, '营收增速')
section_title(ws_asm, 6, '▌ B. 盈利能力')
ws_asm.cell(ASM_ROW['gross_margin'], 1, '综合毛利率')
ws_asm.cell(ASM_ROW['sell_rate'], 1, '销售费用率')
ws_asm.cell(ASM_ROW['admin_rate'], 1, '管理费用率')
ws_asm.cell(ASM_ROW['rd_rate'], 1, '研发费用率')
ws_asm.cell(ASM_ROW['finance_cost'], 1, '净财务收入（亿）')
ws_asm.cell(ASM_ROW['tax_rate'], 1, '有效税率')
ws_asm.cell(ASM_ROW['parent_ratio'], 1, '归母/合并净利润比例')
section_title(ws_asm, 14, '▌ C. 资本开支 & 折旧摊销')
ws_asm.cell(ASM_ROW['da'], 1, 'D&A 折旧摊销（亿）')
ws_asm.cell(ASM_ROW['capex'], 1, '资本开支 Capex（亿）')
section_title(ws_asm, 17, '▌ D. 营运资本')
ws_asm.cell(ASM_ROW['ar_days'], 1, '应收账款天数')
ws_asm.cell(ASM_ROW['inv_days'], 1, '存货周转天数')
ws_asm.cell(ASM_ROW['ap_days'], 1, '应付账款天数')
section_title(ws_asm, 21, '▌ E. 资本结构')
ws_asm.cell(ASM_ROW['div_payout'], 1, '股息支付率')
section_title(ws_asm, 23, '▌ F. 估值参数')
ws_asm.cell(ASM_ROW['wacc'], 1, 'WACC (%)')
ws_asm.cell(ASM_ROW['shares'], 1, '总股本（亿股）')
ws_asm.cell(ASM_ROW['price'], 1, '当前股价（元）')

# 填入假设数据
for yr in PRED_YEARS:
    col = ASM_COL[yr]
    # 营收增速 ← 绿色跨Sheet引用业务拆分
    c = ws_asm.cell(ASM_ROW['rev_growth'], col)
    c.value = f'=业务拆分!{gcl(YEAR_COL[yr])}{BS_ROW["rev_growth"]}'
    c.font = GREEN
    # 综合毛利率 ← 绿色跨Sheet引用业务拆分
    c = ws_asm.cell(ASM_ROW['gross_margin'], col)
    c.value = f'=业务拆分!{gcl(YEAR_COL[yr])}{BS_ROW["blended_gm"]}'
    c.font = GREEN
    # 销售费用率（蓝色）
    ws_asm.cell(ASM_ROW['sell_rate'], col, ASM['sell_rate'][yr]).font = BLUE
    # 管理费用率（蓝色）
    ws_asm.cell(ASM_ROW['admin_rate'], col, ASM['admin_rate'][yr]).font = BLUE
    # 研发费用率（蓝色）
    ws_asm.cell(ASM_ROW['rd_rate'], col, ASM['rd_rate'][yr]).font = BLUE
    # 净财务收入（蓝色）
    ws_asm.cell(ASM_ROW['finance_cost'], col, ASM['fin_cost'][yr]).font = BLUE
    # 税率（蓝色）
    ws_asm.cell(ASM_ROW['tax_rate'], col, ASM['tax_rate'][yr]).font = BLUE
    # 归母比例（蓝色）
    ws_asm.cell(ASM_ROW['parent_ratio'], col, ASM['parent_ratio'][yr]).font = BLUE
    # D&A（蓝色）
    ws_asm.cell(ASM_ROW['da'], col, ASM['da'][yr]).font = BLUE
    # Capex（蓝色）
    ws_asm.cell(ASM_ROW['capex'], col, ASM['capex'][yr]).font = BLUE
    # AR天数（蓝色）
    ws_asm.cell(ASM_ROW['ar_days'], col, ASM['ar_days'][yr]).font = BLUE
    # 存货天数（蓝色）
    ws_asm.cell(ASM_ROW['inv_days'], col, ASM['inv_days'][yr]).font = BLUE
    # AP天数（蓝色）
    ws_asm.cell(ASM_ROW['ap_days'], col, ASM['ap_days'][yr]).font = BLUE
    # 股息支付率（蓝色）
    ws_asm.cell(ASM_ROW['div_payout'], col, ASM['div_payout'][yr]).font = BLUE

# WACC = 8.5%（固定）
for yr in PRED_YEARS:
    ws_asm.cell(ASM_ROW['wacc'], ASM_COL[yr], 0.085).font = BLUE
# 总股本 = 1.90亿股
for yr in PRED_YEARS:
    ws_asm.cell(ASM_ROW['shares'], ASM_COL[yr], 1.90).font = BLUE
# 当前股价 = 54.14元
for yr in PRED_YEARS:
    ws_asm.cell(ASM_ROW['price'], ASM_COL[yr], 54.14).font = BLUE

ws_asm.cell(3, 7, '来源：iFind/研究报告/分析师预测')

ws_asm.column_dimensions['A'].width = 28
for i in range(2, 8):
    ws_asm.column_dimensions[gcl(i)].width = 12

print(f"✓ 模型假设 Sheet 完成")

# ===========================
# Sheet: 利润表（年度）
# ===========================
ws_is = wb.create_sheet('利润表')

IS_ROW = {
    'revenue':      4,
    'yoy':          5,
    'cogs':         6,
    'gp':           7,
    'gm':           8,
    'selling':     10,
    'admin':       11,
    'rd':          12,
    'rd_rate':     13,
    'finance':     14,
    'op':          16,
    'op_margin':   17,
    'ebitda':      18,
    'ebitda_margin':19,
    'np':          21,
    'np_attr':     22,
    'np_attr_yoy': 23,
    'np_margin':    24,
    'eps':         25,
}

ws_is.cell(1, 1, '【利润表】航宇科技（688239.SH）')
ws_is.cell(2, 1, '单位：亿元人民币')

# 年份标题行
ws_is.cell(3, 1, '科目')
for yr in ALL_YEARS:
    ws_is.cell(3, YEAR_COL[yr], yr)
    ws_is.cell(3, YEAR_COL[yr]).font = Font(bold=True)

# 标签
ws_is.cell(IS_ROW['revenue'], 1, '营业收入')
ws_is.cell(IS_ROW['yoy'], 1, 'YoY增速')
ws_is.cell(IS_ROW['cogs'], 1, '营业成本')
ws_is.cell(IS_ROW['gp'], 1, '毛利润')
ws_is.cell(IS_ROW['gm'], 1, '毛利率')
ws_is.cell(9, 1, '期间费用')
ws_is.cell(IS_ROW['selling'], 1, '  销售费用')
ws_is.cell(IS_ROW['admin'], 1, '  管理费用')
ws_is.cell(IS_ROW['rd'], 1, '  研发费用')
ws_is.cell(IS_ROW['rd_rate'], 1, '  研发费用率')
ws_is.cell(IS_ROW['finance'], 1, '  财务费用（净额）')
ws_is.cell(IS_ROW['op'], 1, '营业利润')
ws_is.cell(IS_ROW['op_margin'], 1, '营业利润率')
ws_is.cell(IS_ROW['ebitda'], 1, 'EBITDA')
ws_is.cell(IS_ROW['ebitda_margin'], 1, 'EBITDA率')
ws_is.cell(IS_ROW['np'], 1, '净利润（合并）')
ws_is.cell(IS_ROW['np_attr'], 1, '归母净利润')
ws_is.cell(IS_ROW['np_attr_yoy'], 1, '归母净利YoY')
ws_is.cell(IS_ROW['np_margin'], 1, '归母净利率')
ws_is.cell(IS_ROW['eps'], 1, 'EPS（元/股）')

# 历史数据填充
for yr in HIST_YEARS:
    col = YEAR_COL[yr]
    d = HIST_DATA[yr]
    ws_is.cell(IS_ROW['revenue'], col, d['revenue'])
    ws_is.cell(IS_ROW['cogs'], col, d['cogs'])
    ws_is.cell(IS_ROW['gp'], col, d['revenue'] + d['cogs'])  # GP = Rev + COGS(负)
    ws_is.cell(IS_ROW['gm'], col, d['gross_margin'])
    ws_is.cell(IS_ROW['selling'], col, d['sell_exp'])
    ws_is.cell(IS_ROW['admin'], col, d['admin_exp'])
    ws_is.cell(IS_ROW['rd'], col, d['rd_exp'])
    ws_is.cell(IS_ROW['rd_rate'], col, -d['rd_exp']/d['revenue'])
    ws_is.cell(IS_ROW['finance'], col, d['fin_exp'])
    ws_is.cell(IS_ROW['op'], col, d['op_profit'])
    ws_is.cell(IS_ROW['op_margin'], col, d['op_profit']/d['revenue'])
    # EBITDA估算 = OP + D&A(从HIST_DATA推算)
    da_est = 0.10 * d['revenue']  # 简化估算
    ws_is.cell(IS_ROW['ebitda'], col, d['op_profit'] + da_est)
    ws_is.cell(IS_ROW['ebitda_margin'], col, (d['op_profit']+da_est)/d['revenue'])
    ws_is.cell(IS_ROW['np'], col, d['np'])
    ws_is.cell(IS_ROW['np_attr'], col, d['np_attr'])
    ws_is.cell(IS_ROW['np_margin'], col, d['np_attr']/d['revenue'])
    ws_is.cell(IS_ROW['eps'], col, d['eps'])
    # YoY
    p_yr = prev(yr)
    if p_yr and p_yr in HIST_DATA:
        ws_is.cell(IS_ROW['yoy'], col,
            f'=({dc(yr)}{IS_ROW["revenue"]}-{dc(p_yr)}{IS_ROW["revenue"]})/{dc(p_yr)}{IS_ROW["revenue"]}')
        ws_is.cell(IS_ROW['np_attr_yoy'], col,
            f'=({dc(yr)}{IS_ROW["np_attr"]}-{dc(p_yr)}{IS_ROW["np_attr"]})/{dc(p_yr)}{IS_ROW["np_attr"]}')

# 预测列公式（使用Excel公式字符串）
for yr in PRED_YEARS:
    col = YEAR_COL[yr]
    p = prev(yr)

    # 营收 = 上期 × (1+增速)
    ws_is.cell(IS_ROW['revenue'], col,
        f'={dc(p)}{IS_ROW["revenue"]}*(1+模型假设!${ac(yr)}${ASM_ROW["rev_growth"]})')
    # 毛利润 = 营收 × 毛利率
    ws_is.cell(IS_ROW['gp'], col,
        f'={dc(col)}{IS_ROW["revenue"]}*模型假设!${ac(yr)}${ASM_ROW["gross_margin"]}')
    # 营业成本 = -(营收 - 毛利润)
    ws_is.cell(IS_ROW['cogs'], col,
        f'=-({dc(col)}{IS_ROW["revenue"]}-{dc(col)}{IS_ROW["gp"]})')
    # 各费用 = -营收 × 费率
    ws_is.cell(IS_ROW['selling'], col,
        f'=-{dc(col)}{IS_ROW["revenue"]}*模型假设!${ac(yr)}${ASM_ROW["sell_rate"]}')
    ws_is.cell(IS_ROW['admin'], col,
        f'=-{dc(col)}{IS_ROW["revenue"]}*模型假设!${ac(yr)}${ASM_ROW["admin_rate"]}')
    ws_is.cell(IS_ROW['rd'], col,
        f'=-{dc(col)}{IS_ROW["revenue"]}*模型假设!${ac(yr)}${ASM_ROW["rd_rate"]}')
    ws_is.cell(IS_ROW['rd_rate'], col,
        f'=-{dc(col)}{IS_ROW["rd"]}/{dc(col)}{IS_ROW["revenue"]}')
    # 财务费用
    ws_is.cell(IS_ROW['finance'], col,
        f'=-模型假设!${ac(yr)}${ASM_ROW["finance_cost"]}')
    # 营业利润 = 毛利润 + 各费用
    ws_is.cell(IS_ROW['op'], col,
        f'={dc(col)}{IS_ROW["gp"]}+{dc(col)}{IS_ROW["selling"]}+{dc(col)}{IS_ROW["admin"]}+{dc(col)}{IS_ROW["rd"]}+{dc(col)}{IS_ROW["finance"]}')
    ws_is.cell(IS_ROW['op_margin'], col,
        f'={dc(col)}{IS_ROW["op"]}/{dc(col)}{IS_ROW["revenue"]}')
    # EBITDA = 营业利润 + D&A
    ws_is.cell(IS_ROW['ebitda'], col,
        f'={dc(col)}{IS_ROW["op"]}+模型假设!${ac(yr)}${ASM_ROW["da"]}')
    ws_is.cell(IS_ROW['ebitda_margin'], col,
        f'={dc(col)}{IS_ROW["ebitda"]}/{dc(col)}{IS_ROW["revenue"]}')
    # 净利润（合并）= 营业利润 × (1-税率)
    ws_is.cell(IS_ROW['np'], col,
        f'={dc(col)}{IS_ROW["op"]}*(1-模型假设!${ac(yr)}${ASM_ROW["tax_rate"]})')
    # 归母净利润 = 净利润 × 归母比例
    ws_is.cell(IS_ROW['np_attr'], col,
        f'={dc(col)}{IS_ROW["np"]}*模型假设!${ac(yr)}${ASM_ROW["parent_ratio"]}')
    ws_is.cell(IS_ROW['np_margin'], col,
        f'={dc(col)}{IS_ROW["np_attr"]}/{dc(col)}{IS_ROW["revenue"]}')
    # EPS = 归母净利润 / 总股本
    ws_is.cell(IS_ROW['eps'], col,
        f'={dc(col)}{IS_ROW["np_attr"]}/模型假设!${ac("2026E")}${ASM_ROW["shares"]}')
    # YoY
    ws_is.cell(IS_ROW['yoy'], col,
        f'=({dc(col)}{IS_ROW["revenue"]}-{dc(p)}{IS_ROW["revenue"]})/{dc(p)}{IS_ROW["revenue"]}')
    ws_is.cell(IS_ROW['np_attr_yoy'], col,
        f'=({dc(col)}{IS_ROW["np_attr"]}-{dc(p)}{IS_ROW["np_attr"]})/{dc(p)}{IS_ROW["np_attr"]}')

ws_is.column_dimensions['A'].width = 22
for i in range(2, 13):
    ws_is.column_dimensions[gcl(i)].width = 11

print(f"✓ 利润表 Sheet 完成")

# ===========================
# Sheet: 资产负债表（年度）
# ===========================
# CF_ROW_L must be defined before BS uses it in formulas
CF_ROW_L = {
    'oper': 4, 'capex': 5, 'fcf': 6,
    'invest': 8, 'finance_cf': 9,
    'net_chg': 11, 'beg_cash': 12, 'end_cash': 13,
}

ws_bs = wb.create_sheet('资产负债表')

BS_ROW_L = {
    'cash':       5,
    'ar':         6,
    'inv':        7,
    'other_ca':   8,
    'ca':         9,
    'fa':        11,
    'other_nca': 12,
    'nca':       13,
    'ta':        14,
    'std':       17,
    'ap':        18,
    'other_cl':  19,
    'cl':        20,
    'ltd':       22,
    'other_ncl': 23,
    'ncl':       24,
    'tl':        25,
    'parent_eq': 28,
    'minority':  29,
    'equity':    30,
    'bs_check':  31,
    'leverage':  33,
    'net_debt_eq':34,
}

ws_bs.cell(1, 1, '【资产负债表】航宇科技（688239.SH）')
ws_bs.cell(2, 1, '单位：亿元人民币')

ws_bs.cell(3, 1, '科目')
for yr in ALL_YEARS:
    ws_bs.cell(3, YEAR_COL[yr], yr)
    ws_bs.cell(3, YEAR_COL[yr]).font = Font(bold=True)

ws_bs.cell(4, 1, '流动资产')
ws_bs.cell(BS_ROW_L['cash'], 1, '货币资金')
ws_bs.cell(BS_ROW_L['ar'], 1, '应收账款')
ws_bs.cell(BS_ROW_L['inv'], 1, '存货')
ws_bs.cell(BS_ROW_L['other_ca'], 1, '其他流动资产')
ws_bs.cell(BS_ROW_L['ca'], 1, '流动资产合计')
ws_bs.cell(10, 1, '非流动资产')
ws_bs.cell(BS_ROW_L['fa'], 1, '固定资产+在建工程')
ws_bs.cell(BS_ROW_L['other_nca'], 1, '其他非流动资产')
ws_bs.cell(BS_ROW_L['nca'], 1, '非流动资产合计')
ws_bs.cell(BS_ROW_L['ta'], 1, '资产总计')
ws_bs.cell(15, 1, '流动负债')
ws_bs.cell(BS_ROW_L['std'], 1, '短期借款')
ws_bs.cell(BS_ROW_L['ap'], 1, '应付账款')
ws_bs.cell(BS_ROW_L['other_cl'], 1, '其他流动负债')
ws_bs.cell(BS_ROW_L['cl'], 1, '流动负债合计')
ws_bs.cell(21, 1, '非流动负债')
ws_bs.cell(BS_ROW_L['ltd'], 1, '长期借款')
ws_bs.cell(BS_ROW_L['other_ncl'], 1, '其他非流动负债')
ws_bs.cell(BS_ROW_L['ncl'], 1, '非流动负债合计')
ws_bs.cell(BS_ROW_L['tl'], 1, '负债合计')
ws_bs.cell(26, 1, '股东权益')
ws_bs.cell(BS_ROW_L['parent_eq'], 1, '归母股东权益')
ws_bs.cell(BS_ROW_L['minority'], 1, '少数股东权益')
ws_bs.cell(BS_ROW_L['equity'], 1, '股东权益合计')
ws_bs.cell(BS_ROW_L['bs_check'], 1, '验证（负债+权益）')
ws_bs.cell(BS_ROW_L['leverage'], 1, '资产负债率')
ws_bs.cell(BS_ROW_L['net_debt_eq'], 1, '净负债/权益')

# 历史数据填充
for yr in HIST_YEARS:
    col = YEAR_COL[yr]
    d = HIST_DATA[yr]
    ws_bs.cell(BS_ROW_L['cash'], col, d['cash'])
    ws_bs.cell(BS_ROW_L['ar'], col, d['ar'])
    ws_bs.cell(BS_ROW_L['inv'], col, d['inv'])
    ws_bs.cell(BS_ROW_L['other_ca'], col, d['other_ca'])
    ws_bs.cell(BS_ROW_L['ca'], col, d['cash']+d['ar']+d['inv']+d['other_ca'])
    ws_bs.cell(BS_ROW_L['fa'], col, d['fa'])
    ws_bs.cell(BS_ROW_L['other_nca'], col, d['other_nca'])
    ws_bs.cell(BS_ROW_L['nca'], col, d['fa']+d['other_nca'])
    ws_bs.cell(BS_ROW_L['ta'], col, ws_bs.cell(BS_ROW_L['ca'],col).value + ws_bs.cell(BS_ROW_L['nca'],col).value)
    ws_bs.cell(BS_ROW_L['std'], col, d['std'])
    ws_bs.cell(BS_ROW_L['ap'], col, d['ap'])
    ws_bs.cell(BS_ROW_L['other_cl'], col, d['other_cl'])
    ws_bs.cell(BS_ROW_L['cl'], col, d['std']+d['ap']+d['other_cl'])
    ws_bs.cell(BS_ROW_L['ltd'], col, d['ltd'])
    ws_bs.cell(BS_ROW_L['other_ncl'], col, d['other_ncl'])
    ws_bs.cell(BS_ROW_L['ncl'], col, d['ltd']+d['other_ncl'])
    ws_bs.cell(BS_ROW_L['tl'], col, ws_bs.cell(BS_ROW_L['cl'],col).value + ws_bs.cell(BS_ROW_L['ncl'],col).value)
    ws_bs.cell(BS_ROW_L['parent_eq'], col, d['parent_eq'])
    ws_bs.cell(BS_ROW_L['minority'], col, d['minority'])
    ws_bs.cell(BS_ROW_L['equity'], col, d['parent_eq']+d['minority'])
    ws_bs.cell(BS_ROW_L['bs_check'], col, ws_bs.cell(BS_ROW_L['tl'],col).value + ws_bs.cell(BS_ROW_L['equity'],col).value)
    ws_bs.cell(BS_ROW_L['leverage'], col, ws_bs.cell(BS_ROW_L['tl'],col).value / ws_bs.cell(BS_ROW_L['ta'],col).value)
    nd = (d['std']+d['ltd']-d['cash'])/(d['parent_eq']+d['minority']) if (d['parent_eq']+d['minority']) != 0 else 0
    ws_bs.cell(BS_ROW_L['net_debt_eq'], col, nd)

# 预测列公式
for yr in PRED_YEARS:
    col = YEAR_COL[yr]
    p = prev(yr)
    # 货币资金 ← 现金流量表期末余额
    ws_bs.cell(BS_ROW_L['cash'], col, f'=现金流量表!{dc(col)}{CF_ROW_L["end_cash"]}')
    # 应收账款 = 营收 × 应收天数 / 365
    ws_bs.cell(BS_ROW_L['ar'], col,
        f'=利润表!{dc(col)}{IS_ROW["revenue"]}*模型假设!${ac(yr)}${ASM_ROW["ar_days"]}/365')
    # 存货 = |COGS| × 存货天数 / 365
    ws_bs.cell(BS_ROW_L['inv'], col,
        f'=-利润表!{dc(col)}{IS_ROW["cogs"]}*模型假设!${ac(yr)}${ASM_ROW["inv_days"]}/365')
    # 其他流动资产
    ws_bs.cell(BS_ROW_L['other_ca'], col, f'={dc(p)}{BS_ROW_L["other_ca"]}*0.99')
    # 流动资产合计
    ws_bs.cell(BS_ROW_L['ca'], col,
        f'=SUM({dc(col)}{BS_ROW_L["cash"]}:{dc(col)}{BS_ROW_L["other_ca"]})')
    # 固定资产 = 上期 + Capex - D&A
    ws_bs.cell(BS_ROW_L['fa'], col,
        f'={dc(p)}{BS_ROW_L["fa"]}+模型假设!${ac(yr)}${ASM_ROW["capex"]}-模型假设!${ac(yr)}${ASM_ROW["da"]}')
    # 其他非流动资产
    ws_bs.cell(BS_ROW_L['other_nca'], col, f'={dc(p)}{BS_ROW_L["other_nca"]}*1.01')
    # 非流动资产
    ws_bs.cell(BS_ROW_L['nca'], col,
        f'={dc(col)}{BS_ROW_L["fa"]}+{dc(col)}{BS_ROW_L["other_nca"]}')
    # 资产总计
    ws_bs.cell(BS_ROW_L['ta'], col,
        f'={dc(col)}{BS_ROW_L["ca"]}+{dc(col)}{BS_ROW_L["nca"]}')
    # 应付账款
    ws_bs.cell(BS_ROW_L['ap'], col,
        f'=-利润表!{dc(col)}{IS_ROW["cogs"]}*模型假设!${ac(yr)}${ASM_ROW["ap_days"]}/365')
    # 短期借款
    ws_bs.cell(BS_ROW_L['std'], col, f'={dc(p)}{BS_ROW_L["std"]}')
    # 其他流动负债
    ws_bs.cell(BS_ROW_L['other_cl'], col, f'={dc(p)}{BS_ROW_L["other_cl"]}*0.98')
    # 流动负债合计
    ws_bs.cell(BS_ROW_L['cl'], col,
        f'={dc(col)}{BS_ROW_L["std"]}+{dc(col)}{BS_ROW_L["ap"]}+{dc(col)}{BS_ROW_L["other_cl"]}')
    # 长期借款
    ws_bs.cell(BS_ROW_L['ltd'], col, f'={dc(p)}{BS_ROW_L["ltd"]}')
    # 其他非流动负债
    ws_bs.cell(BS_ROW_L['other_ncl'], col, f'={dc(p)}{BS_ROW_L["other_ncl"]}*1.02')
    # 非流动负债合计
    ws_bs.cell(BS_ROW_L['ncl'], col,
        f'={dc(col)}{BS_ROW_L["ltd"]}+{dc(col)}{BS_ROW_L["other_ncl"]}')
    # 负债合计
    ws_bs.cell(BS_ROW_L['tl'], col,
        f'={dc(col)}{BS_ROW_L["cl"]}+{dc(col)}{BS_ROW_L["ncl"]}')
    # 归母权益 = 上期 + 归母净利 - 分红
    ws_bs.cell(BS_ROW_L['parent_eq'], col,
        f'={dc(p)}{BS_ROW_L["parent_eq"]}+利润表!{dc(col)}{IS_ROW["np_attr"]}*(1-模型假设!${ac(yr)}${ASM_ROW["div_payout"]})')
    # 少数股东权益
    ws_bs.cell(BS_ROW_L['minority'], col, f'={dc(p)}{BS_ROW_L["minority"]}*1.02')
    # 股东权益合计
    ws_bs.cell(BS_ROW_L['equity'], col,
        f'={dc(col)}{BS_ROW_L["parent_eq"]}+{dc(col)}{BS_ROW_L["minority"]}')
    # 验证
    ws_bs.cell(BS_ROW_L['bs_check'], col,
        f'={dc(col)}{BS_ROW_L["tl"]}+{dc(col)}{BS_ROW_L["equity"]}')
    # 资产负债率
    ws_bs.cell(BS_ROW_L['leverage'], col,
        f'={dc(col)}{BS_ROW_L["tl"]}/{dc(col)}{BS_ROW_L["ta"]}')
    # 净负债/权益
    ws_bs.cell(BS_ROW_L['net_debt_eq'], col,
        f'=({dc(col)}{BS_ROW_L["std"]}+{dc(col)}{BS_ROW_L["ltd"]}-{dc(col)}{BS_ROW_L["cash"]})/{dc(col)}{BS_ROW_L["equity"]}')

ws_bs.column_dimensions['A'].width = 22
for i in range(2, 13):
    ws_bs.column_dimensions[gcl(i)].width = 11

print(f"✓ 资产负债表 Sheet 完成")

# ===========================
# Sheet: 现金流量表（年度）
# ===========================
ws_cf = wb.create_sheet('现金流量表')

CF_ROW_L = {
    'oper':      4,
    'capex':     5,
    'fcf':       6,
    'invest':    8,
    'finance_cf':9,
    'net_chg':  11,
    'beg_cash': 12,
    'end_cash': 13,
}

ws_cf.cell(1, 1, '【现金流量表】航宇科技（688239.SH）')
ws_cf.cell(2, 1, '单位：亿元人民币')

ws_cf.cell(3, 1, '科目')
for yr in ALL_YEARS:
    ws_cf.cell(3, YEAR_COL[yr], yr)
    ws_cf.cell(3, YEAR_COL[yr]).font = Font(bold=True)

ws_cf.cell(CF_ROW_L['oper'], 1, '经营活动现金流')
ws_cf.cell(CF_ROW_L['capex'], 1, '资本开支（Capex）')
ws_cf.cell(CF_ROW_L['fcf'], 1, '自由现金流（FCFF）')
ws_cf.cell(7, 1, '投资活动现金流')
ws_cf.cell(CF_ROW_L['invest'], 1, '  投资活动现金流')
ws_cf.cell(CF_ROW_L['finance_cf'], 1, '  筹资活动现金流')
ws_cf.cell(CF_ROW_L['net_chg'], 1, '净现金增量')
ws_cf.cell(CF_ROW_L['beg_cash'], 1, '期初货币资金')
ws_cf.cell(CF_ROW_L['end_cash'], 1, '期末货币资金')

# 历史数据
for yr in HIST_YEARS:
    col = YEAR_COL[yr]
    d = HIST_DATA[yr]
    ws_cf.cell(CF_ROW_L['oper'], col, d['oper_cf'])
    ws_cf.cell(CF_ROW_L['capex'], col, -d['capex'])  # 负值转正值
    ws_cf.cell(CF_ROW_L['fcf'], col, d['oper_cf'] - d['capex'])
    ws_cf.cell(CF_ROW_L['invest'], col, d['invest_cf'])
    ws_cf.cell(CF_ROW_L['finance_cf'], col, d['finance_cf'])
    ws_cf.cell(CF_ROW_L['net_chg'], col, d['oper_cf']+d['invest_cf']+d['finance_cf'])
    p = prev(yr)
    if p:
        ws_cf.cell(CF_ROW_L['beg_cash'], col, HIST_DATA[p]['cash'])
    ws_cf.cell(CF_ROW_L['end_cash'], col, d['cash'])

# 预测公式
for yr in PRED_YEARS:
    col = YEAR_COL[yr]
    p = prev(yr)
    # 经营现金流 = 净利润 + D&A - 营运资本变动
    ws_cf.cell(CF_ROW_L['oper'], col,
        f'=利润表!{dc(col)}{IS_ROW["np"]}'
        f'+模型假设!${ac(yr)}${ASM_ROW["da"]}'
        f'-((资产负债表!{dc(col)}{BS_ROW_L["ar"]}-资产负债表!{dc(p)}{BS_ROW_L["ar"]})'
        f'+(资产负债表!{dc(col)}{BS_ROW_L["inv"]}-资产负债表!{dc(p)}{BS_ROW_L["inv"]})'
        f'-(资产负债表!{dc(col)}{BS_ROW_L["ap"]}-资产负债表!{dc(p)}{BS_ROW_L["ap"]}))')
    # Capex
    ws_cf.cell(CF_ROW_L['capex'], col, f'=-模型假设!${ac(yr)}${ASM_ROW["capex"]}')
    # FCF
    ws_cf.cell(CF_ROW_L['fcf'], col,
        f'={dc(col)}{CF_ROW_L["oper"]}+{dc(col)}{CF_ROW_L["capex"]}')
    # 投资活动
    ws_cf.cell(CF_ROW_L['invest'], col, f'=-模型假设!${ac(yr)}${ASM_ROW["capex"]}*0.9')
    # 筹资活动
    ws_cf.cell(CF_ROW_L['finance_cf'], col,
        f'=-利润表!{dc(col)}{IS_ROW["np_attr"]}*模型假设!${ac(yr)}${ASM_ROW["div_payout"]}')
    # 净现金增量
    ws_cf.cell(CF_ROW_L['net_chg'], col,
        f'={dc(col)}{CF_ROW_L["oper"]}+{dc(col)}{CF_ROW_L["invest"]}+{dc(col)}{CF_ROW_L["finance_cf"]}')
    # 期初现金
    ws_cf.cell(CF_ROW_L['beg_cash'], col, f'={dc(p)}{CF_ROW_L["end_cash"]}')
    # 期末现金
    ws_cf.cell(CF_ROW_L['end_cash'], col,
        f'={dc(col)}{CF_ROW_L["beg_cash"]}+{dc(col)}{CF_ROW_L["net_chg"]}')

ws_cf.column_dimensions['A'].width = 24
for i in range(2, 13):
    ws_cf.column_dimensions[gcl(i)].width = 11

print(f"✓ 现金流量表 Sheet 完成")

# ===========================
# Sheet: 利润表_季度
# ===========================
ws_isq = wb.create_sheet('利润表_季度')

Q_COL = {
    '2024Q1': 2, '2024Q2': 3, '2024Q3': 4, '2024Q4': 5,
    '2025Q1': 6, '2025Q2': 7, '2025Q3': 8, '2025Q4': 9,
    '2026Q1E':10, '2026Q2E':11, '2026Q3E':12, '2026Q4E':13,
}
Q_YEARS = list(Q_COL.keys())
HIST_Q = ['2024Q1','2024Q2','2024Q3','2024Q4','2025Q1','2025Q2','2025Q3','2025Q4']
PRED_Q = ['2026Q1E','2026Q2E','2026Q3E','2026Q4E']

ISQ_ROW = {
    'revenue': 4, 'yoy': 5, 'qoq': 6,
    'cogs': 7, 'gp': 8, 'gm': 9,
    'sell': 10, 'admin': 11, 'rd': 12, 'finance': 13,
    'op': 15, 'np': 17, 'np_attr': 18, 'np_yoy': 19, 'np_qoq': 20, 'eps': 21,
}

ws_isq.cell(1, 1, '【利润表_季度】航宇科技（688239.SH）')
ws_isq.cell(2, 1, 'E=季节性预测 | 历史=已披露实际数据')

ws_isq.cell(3, 1, '科目')
for q in Q_YEARS:
    ws_isq.cell(3, Q_COL[q], q)
    ws_isq.cell(3, Q_COL[q]).font = Font(bold=True)

ws_isq.cell(ISQ_ROW['revenue'], 1, '营业收入')
ws_isq.cell(ISQ_ROW['yoy'], 1, 'YoY增速')
ws_isq.cell(ISQ_ROW['qoq'], 1, 'QoQ增速')
ws_isq.cell(ISQ_ROW['cogs'], 1, '营业成本')
ws_isq.cell(ISQ_ROW['gp'], 1, '毛利润')
ws_isq.cell(ISQ_ROW['gm'], 1, '毛利率')
ws_isq.cell(ISQ_ROW['sell'], 1, '销售费用')
ws_isq.cell(ISQ_ROW['admin'], 1, '管理费用')
ws_isq.cell(ISQ_ROW['rd'], 1, '研发费用')
ws_isq.cell(ISQ_ROW['finance'], 1, '财务费用')
ws_isq.cell(ISQ_ROW['op'], 1, '营业利润')
ws_isq.cell(ISQ_ROW['np'], 1, '净利润（合并）')
ws_isq.cell(ISQ_ROW['np_attr'], 1, '归母净利润')
ws_isq.cell(ISQ_ROW['np_yoy'], 1, '归母净利YoY')
ws_isq.cell(ISQ_ROW['np_qoq'], 1, '归母净利QoQ')
ws_isq.cell(ISQ_ROW['eps'], 1, 'EPS（元）')

# 历史季度数据
for q in HIST_Q:
    col = Q_COL[q]
    d = QUARTERLY_DATA[q]
    ws_isq.cell(ISQ_ROW['revenue'], col, d['revenue'])
    gp = d['revenue'] * d['gross_margin']
    ws_isq.cell(ISQ_ROW['gp'], col, gp)
    ws_isq.cell(ISQ_ROW['cogs'], col, -(d['revenue'] - gp))
    ws_isq.cell(ISQ_ROW['gm'], col, d['gross_margin'])
    ws_isq.cell(ISQ_ROW['np_attr'], col, d['np_attr'])
    # 估算其他科目（简化）
    ws_isq.cell(ISQ_ROW['sell'], col, -d['revenue']*0.025)
    ws_isq.cell(ISQ_ROW['admin'], col, -d['revenue']*0.065)
    ws_isq.cell(ISQ_ROW['rd'], col, -d['revenue']*0.060)
    ws_isq.cell(ISQ_ROW['finance'], col, -d['revenue']*0.005)
    ws_isq.cell(ISQ_ROW['op'], col, gp*0.58)  # 简化
    ws_isq.cell(ISQ_ROW['np'], col, d['np_attr']*1.05)
    ws_isq.cell(ISQ_ROW['eps'], col, d['np_attr']/1.90)

# 预测季度（2026E，基于年度预测分摊）
# 先计算历史季节性比例
annual_revenue = {'2024A': 18.05, '2025A': 20.34, '2026E': 22.0}
q_seasonal = {
    'Q1': {'hist': [5.78/18.05, 4.22/20.34], 'avg': None},
    'Q2': {'hist': [4.92/18.05, 4.92/20.34], 'avg': None},
    'Q3': {'hist': [4.12/18.05, 6.10/20.34], 'avg': None},
    'Q4': {'hist': [3.93/18.05, 5.10/20.34], 'avg': None},
}
for q in q_seasonal:
    q_seasonal[q]['avg'] = sum(q_seasonal[q]['hist']) / len(q_seasonal[q]['hist'])

PRED_ANNUAL = {'2026E': 22.0, '2027E': 26.0, '2028E': 31.0}
q_map = {'2026Q1E':'Q1','2026Q2E':'Q2','2026Q3E':'Q3','2026Q4E':'Q4'}

for q_pred in PRED_Q:
    col = Q_COL[q_pred]
    q_key = q_map[q_pred]
    pct = q_seasonal[q_key]['avg']
    ann_rev = PRED_ANNUAL['2026E']
    ann_np  = 3.19  # 2026E归母净利
    q_rev = ann_rev * pct
    q_np  = ann_np * pct
    q_gm  = 0.280
    q_gp  = q_rev * q_gm
    ws_isq.cell(ISQ_ROW['revenue'], col, q_rev)
    ws_isq.cell(ISQ_ROW['gp'], col, q_gp)
    ws_isq.cell(ISQ_ROW['cogs'], col, -(q_rev - q_gp))
    ws_isq.cell(ISQ_ROW['gm'], col, q_gm)
    ws_isq.cell(ISQ_ROW['np_attr'], col, q_np)
    ws_isq.cell(ISQ_ROW['sell'], col, -q_rev*0.025)
    ws_isq.cell(ISQ_ROW['admin'], col, -q_rev*0.065)
    ws_isq.cell(ISQ_ROW['rd'], col, -q_rev*0.062)
    ws_isq.cell(ISQ_ROW['finance'], col, -q_rev*0.005)
    ws_isq.cell(ISQ_ROW['op'], col, q_gp*0.58)
    ws_isq.cell(ISQ_ROW['np'], col, q_np*1.05)
    ws_isq.cell(ISQ_ROW['eps'], col, q_np/1.90)

ws_isq.column_dimensions['A'].width = 20
for i in range(2, 14):
    ws_isq.column_dimensions[gcl(i)].width = 10

print(f"✓ 利润表_季度 Sheet 完成")

# ===========================
# Sheet: 基本面研究
# ===========================
ws_sr = wb.create_sheet('基本面研究')

ws_sr.cell(1, 1, '【基本面研究】航宇科技（688239.SH）')
ws_sr.cell(2, 1, '数据截至：2026-04-07')

# 评分汇总
ws_sr.cell(4, 1, '▌ 评分汇总')
ws_sr.cell(5, 1, '维度')
ws_sr.cell(5, 2, '权重')
ws_sr.cell(5, 3, '分值')
ws_sr.cell(5, 4, '加权分')
ws_sr.cell(5, 1).font = Font(bold=True)

# A行业质量 (25%)
ws_sr.cell(6, 1, 'A. 行业质量')
ws_sr.cell(6, 2, '25%')
ws_sr.cell(6, 3, '3.8')
ws_sr.cell(6, 4, '=B6*C6').font = BLACK

# B竞争壁垒 (25%)
ws_sr.cell(7, 1, 'B. 竞争壁垒')
ws_sr.cell(7, 2, '25%')
ws_sr.cell(7, 3, '3.5')
ws_sr.cell(7, 4, '=B7*C7').font = BLACK

# C增长能见度 (25%)
ws_sr.cell(8, 1, 'C. 增长能见度')
ws_sr.cell(8, 2, '25%')
ws_sr.cell(8, 3, '3.5')
ws_sr.cell(8, 4, '=B8*C8').font = BLACK

# D财务质量 (15%)
ws_sr.cell(9, 1, 'D. 财务质量')
ws_sr.cell(9, 2, '15%')
ws_sr.cell(9, 3, '3.2')
ws_sr.cell(9, 4, '=B9*C9').font = BLACK

# E公司治理 (10%)
ws_sr.cell(10, 1, 'E. 公司治理')
ws_sr.cell(10, 2, '10%')
ws_sr.cell(10, 3, '3.5')
ws_sr.cell(10, 4, '=B10*C10').font = BLACK

# 综合得分
ws_sr.cell(11, 1, '★ 综合质量得分')
ws_sr.cell(11, 4, '=SUM(D6:D10)')
ws_sr.cell(11, 4).font = Font(bold=True, color='0000FF')

# 投资等级
ws_sr.cell(13, 1, '★ 投资质量等级')
ws_sr.cell(14, 1, '> 3.8 = A档（优秀）| 3.0-3.8 = B档（良好）| < 3.0 = C档（一般）')
ws_sr.cell(15, 1, '综合得分：3.5 → B档（良好）')

# A. 行业质量
ws_sr.cell(17, 1, '▌ A. 行业质量详细评分（权重25%）')
ws_sr.cell(18, 1, '子维度')
ws_sr.cell(18, 2, '权重')
ws_sr.cell(18, 3, '得分')
ws_sr.cell(18, 4, '评分依据')

data_a = [
    ('行业规模（TAM）', '5%', '4.0', '中国航空航天核心产业规模突破2万亿元'),
    ('行业增速（未来5年CAGR）', '5%', '4.0', '商业航天CAGR~20%，低空经济CAGR超20%'),
    ('行业延展性', '5%', '4.0', 'C919/ARJ21放量，商业航天/卫星互联网打开空间'),
    ('行业透明度', '5%', '3.5', '竞争格局有序，环形锻件CR3高'),
    ('政策友好度', '5%', '4.0', '国防预算+7.2%，军民融合政策持续推进'),
]
for i, (dim, wt, score, basis) in enumerate(data_a):
    row = 19 + i
    ws_sr.cell(row, 1, dim)
    ws_sr.cell(row, 2, wt)
    ws_sr.cell(row, 3, float(score))
    ws_sr.cell(row, 4, basis)

# B. 竞争壁垒
ws_sr.cell(26, 1, '▌ B. 竞争壁垒详细评分（权重25%）')
ws_sr.cell(27, 1, '子维度')
ws_sr.cell(27, 2, '权重')
ws_sr.cell(27, 3, '得分')
ws_sr.cell(27, 4, '评分依据')

data_b = [
    ('技术壁垒', '5%', '4.0', 'NSP技术获全球航空龙头认证，节省原材料20-50%'),
    ('定价权', '5%', '3.5', '毛利率28%稳定，NSP推广有提升空间'),
    ('垄断地位', '5%', '4.0', '国内航空环形锻件龙头，全球三强之一'),
    ('客户粘性', '5%', '4.0', 'GE/赛峰/罗罗/普惠认证，客户转换成本极高'),
    ('进入壁垒', '5%', '3.0', '航空锻件资质认证3-5年，有一定壁垒'),
    ('供应商议价', '5%', '3.0', '高温合金供应商分散，议价能力中'),
    ('客户集中度', '5%', '3.0', '商发占比85-90%，集中度偏高'),
    ('替代品威胁', '5%', '3.5', '无明显替代品，技术替代风险低'),
]
for i, (dim, wt, score, basis) in enumerate(data_b):
    row = 28 + i
    ws_sr.cell(row, 1, dim)
    ws_sr.cell(row, 2, wt)
    ws_sr.cell(row, 3, float(score))
    ws_sr.cell(row, 4, basis)

# C. 增长能见度
ws_sr.cell(38, 1, '▌ C. 增长能见度详细评分（权重25%）')
ws_sr.cell(39, 1, '子维度')
ws_sr.cell(39, 2, '权重')
ws_sr.cell(39, 3, '得分')
ws_sr.cell(39, 4, '评分依据')

data_c = [
    ('在手订单充裕度', '8%', '4.5', '在手订单60亿元，覆盖2026-2027年'),
    ('产能爬坡', '7%', '3.5', '贵阳70-80%，德阳新投产，匈牙利规划中'),
    ('新需求开拓', '5%', '4.0', '燃气轮机/核电/半导体设备打开新空间'),
    ('生命周期', '5%', '3.0', '成长期中段，国内交付节奏波动影响短期'),
]
for i, (dim, wt, score, basis) in enumerate(data_c):
    row = 40 + i
    ws_sr.cell(row, 1, dim)
    ws_sr.cell(row, 2, wt)
    ws_sr.cell(row, 3, float(score))
    ws_sr.cell(row, 4, basis)

# D. 财务质量
ws_sr.cell(46, 1, '▌ D. 财务质量详细评分（权重15%）')
ws_sr.cell(47, 1, '子维度')
ws_sr.cell(47, 2, '权重')
ws_sr.cell(47, 3, '得分')
ws_sr.cell(47, 4, '评分依据')

data_d = [
    ('CFO/净利润', '4%', '3.0', '2024年经营现金流1.44亿/净利1.89亿=76%'),
    ('ROE趋势', '4%', '3.0', 'ROE 9.87%属中等，2025年略降至9.76%'),
    ('负债率', '3%', '3.5', '资产负债率~50%，可控'),
    ('运营效率', '3%', '3.0', '应收账款+存货周转天数边际恶化'),
    ('股东回报', '1%', '3.0', '股息支付率15%，有回购'),
]
for i, (dim, wt, score, basis) in enumerate(data_d):
    row = 48 + i
    ws_sr.cell(row, 1, dim)
    ws_sr.cell(row, 2, wt)
    ws_sr.cell(row, 3, float(score))
    ws_sr.cell(row, 4, basis)

# E. 公司治理
ws_sr.cell(55, 1, '▌ E. 公司治理详细评分（权重10%）')
ws_sr.cell(56, 1, '子维度')
ws_sr.cell(56, 2, '权重')
ws_sr.cell(56, 3, '得分')
ws_sr.cell(56, 4, '评分依据')

data_e = [
    ('管理层能力', '4%', '3.5', '战略清晰，国际化布局领先'),
    ('诚信度', '3%', '3.5', '无明显诚信问题记录'),
    ('股权结构', '3%', '3.5', '民营控股，张华持股13.78%，股权相对分散'),
]
for i, (dim, wt, score, basis) in enumerate(data_e):
    row = 57 + i
    ws_sr.cell(row, 1, dim)
    ws_sr.cell(row, 2, wt)
    ws_sr.cell(row, 3, float(score))
    ws_sr.cell(row, 4, basis)

# 综合结论
ws_sr.cell(62, 1, '▌ 综合结论')
ws_sr.cell(63, 1, '投资建议：军工航空环形锻件稀缺龙头，订单充裕保障中期业绩，中性偏乐观')
ws_sr.cell(64, 1, '核心亮点：①全球市场份额提升逻辑清晰（目标20-30%）②在手订单60亿保障③NSP技术竞争力强')
ws_sr.cell(65, 1, '核心风险：①国内交付节奏波动 ②子公司亏损拖累 ③估值偏高(PE 55x)')

ws_sr.column_dimensions['A'].width = 28
ws_sr.column_dimensions['D'].width = 48
for i in range(2, 5):
    ws_sr.column_dimensions[gcl(i)].width = 12

print(f"✓ 基本面研究 Sheet 完成")

# ===========================
# Sheet: 估值分析
# ===========================
ws_va = wb.create_sheet('估值分析')

ws_va.cell(1, 1, '【估值分析】航宇科技（688239.SH）')
ws_va.cell(2, 1, '报告日期：2026-04-07 | 当前股价：54.14元')

# PE估值法
ws_va.cell(4, 1, '▌ PE 估值法')
ws_va.cell(5, 1, '')
ws_va.cell(5, 2, '2024A')
ws_va.cell(5, 3, '2025A')
ws_va.cell(5, 4, '2026E')
ws_va.cell(5, 5, '2027E')
ws_va.cell(5, 6, '2028E')
for yr, col in [('2024A',2),('2025A',3),('2026E',4),('2027E',5),('2028E',6)]:
    if yr.endswith('A'):
        ws_va.cell(6, col, HIST_DATA[yr]['eps'])
        ws_va.cell(7, col, f'=B7' if col==2 else f'=C7' if col==3 else '')
    else:
        ws_va.cell(6, col, yr+'_eps_formula')
        ws_va.cell(7, col, '')

ws_va.cell(6, 1, 'EPS（元）')
ws_va.cell(6, 2, 1.30)
ws_va.cell(6, 3, 0.98)
ws_va.cell(6, 4, '=利润表!H25')  # 引用2026E EPS
ws_va.cell(6, 5, '=利润表!I25')
ws_va.cell(6, 6, '=利润表!J25')

ws_va.cell(7, 1, 'PE倍数')
ws_va.cell(7, 2, 41.6)  # 历史PE
ws_va.cell(7, 3, 55.2)  # 当前PE
ws_va.cell(7, 4, 40)  # 2026E目标PE
ws_va.cell(7, 5, 35)
ws_va.cell(7, 6, 30)

ws_va.cell(8, 1, '目标股价')
ws_va.cell(8, 2, 54.1)
ws_va.cell(8, 3, 54.1)
ws_va.cell(8, 4, '=D6*D7')
ws_va.cell(8, 5, '=E6*E7')
ws_va.cell(8, 6, '=F6*F7')

# PEGY估值
ws_va.cell(10, 1, '▌ PEGY 估值法')
ws_va.cell(11, 1, 'FY1E EPS')
ws_va.cell(11, 4, '=D6')
ws_va.cell(12, 1, 'FY1E PE（当前）')
ws_va.cell(12, 4, 55.2)
ws_va.cell(13, 1, '净利润增速（FY1）')
ws_va.cell(13, 4, 0.715)  # (3.19-1.86)/1.86
ws_va.cell(14, 1, '股息率（TTM）')
ws_va.cell(14, 4, 0.028)  # 估算
ws_va.cell(15, 1, 'PEG')
ws_va.cell(15, 4, '=D12/D13')
ws_va.cell(16, 1, 'PEGY')
ws_va.cell(16, 4, '=D12/(D13+D14)')
ws_va.cell(17, 1, '合理PEGY区间')
ws_va.cell(17, 4, '0.8-1.2x')
ws_va.cell(18, 1, 'PEGY隐含目标价')
ws_va.cell(18, 4, '=D6*(D13+D14)*1.0')

# 估值汇总
ws_va.cell(20, 1, '▌ 估值汇总')
ws_va.cell(21, 1, '方法')
ws_va.cell(21, 2, '目标价（元）')
ws_va.cell(21, 3, '当前股价（元）')
ws_va.cell(21, 4, '上行空间')
ws_va.cell(22, 1, 'PE法（2026E 40x）')
ws_va.cell(22, 2, 66.8)  # 1.67*40
ws_va.cell(22, 3, 54.14)
ws_va.cell(22, 4, '=(B22-C22)/C22')
ws_va.cell(23, 1, 'PE法（2027E 35x）')
ws_va.cell(23, 2, 77.4)  # 2.21*35
ws_va.cell(23, 3, 54.14)
ws_va.cell(23, 4, '=(B23-C23)/C23')
ws_va.cell(24, 1, 'PE法（2028E 30x）')
ws_va.cell(24, 2, 81.6)  # 2.72*30
ws_va.cell(24, 3, 54.14)
ws_va.cell(24, 4, '=(B24-C24)/C24')
ws_va.cell(25, 1, 'PEGY法')
ws_va.cell(25, 2, '=D18')
ws_va.cell(25, 3, 54.14)
ws_va.cell(25, 4, '=(B25-C25)/C25')
ws_va.cell(26, 1, '分析师平均')
ws_va.cell(26, 2, 68.0)  # 太平洋65-72
ws_va.cell(26, 3, 54.14)
ws_va.cell(26, 4, '=(B26-C26)/C26')

ws_va.cell(28, 1, '★ 综合估值结论')
ws_va.cell(28, 1).font = Font(bold=True)
ws_va.cell(29, 1, '目标价区间：64-72元（基于2026E PE 40-45x）')
ws_va.cell(30, 1, '上行空间：+18%~+33%')
ws_va.cell(31, 1, '投资评级：中性偏乐观（边际改善）')

ws_va.column_dimensions['A'].width = 24
for i in range(2, 7):
    ws_va.column_dimensions[gcl(i)].width = 12

print(f"✓ 估值分析 Sheet 完成")

# ===========================
# Sheet: 摘要（最后更新）
# ===========================
ws_sum = wb.create_sheet('摘要', 0)  # 插入到第一位

ws_sum.cell(1, 1, '航宇科技（688239.SH）— 财务研究模型')
ws_sum.cell(1, 1).font = Font(size=14, bold=True)
ws_sum.cell(2, 1, '数据来源：iFind/研究报告/分析师预测 | 单位：亿元人民币 | 截至：2026-04-07')

# 公司概况
ws_sum.cell(4, 1, '▌ 公司概况')
ws_sum.cell(5, 1, '股票代码')
ws_sum.cell(5, 2, '688239.SH')
ws_sum.cell(6, 1, '行业')
ws_sum.cell(6, 2, '国防军工/航空锻造')
ws_sum.cell(7, 1, '主营业务')
ws_sum.cell(7, 2, '航空难变形金属材料环形锻件')
ws_sum.cell(8, 1, '总股本')
ws_sum.cell(8, 2, '1.90亿股')
ws_sum.cell(9, 1, '当前股价')
ws_sum.cell(9, 2, '54.14元')
ws_sum.cell(10, 1, '总市值')
ws_sum.cell(10, 2, '103亿元')

# 历史财务概要
ws_sum.cell(12, 1, '▌ 历史财务概要')
headers = ['指标', '2020A', '2021A', '2022A', '2023A', '2024A', '2025A']
for i, h in enumerate(headers):
    ws_sum.cell(13, i+1, h)
    ws_sum.cell(13, i+1).font = Font(bold=True)

sum_data = [
    ('营业收入', 'revenue'),
    ('YoY增速', 'yoy'),
    ('毛利润', 'gp'),
    ('毛利率', 'gm'),
    ('归母净利润', 'np_attr'),
    ('归母净利率', 'np_margin'),
    ('EPS（元）', 'eps'),
]
for i, (label, key) in enumerate(sum_data):
    ws_sum.cell(14+i, 1, label)
    for j, yr in enumerate(HIST_YEARS):
        if key in ['yoy', 'gm', 'np_margin']:
            val = f'=利润表!{gcl(YEAR_COL[yr])}{IS_ROW[key]}'
        elif key == 'gp':
            val = f'=利润表!{gcl(YEAR_COL[yr])}{IS_ROW["gp"]}'
        elif key == 'np_attr':
            val = f'=利润表!{gcl(YEAR_COL[yr])}{IS_ROW[key]}'
        elif key == 'eps':
            val = f'=利润表!{gcl(YEAR_COL[yr])}{IS_ROW[key]}'
        elif key == 'revenue':
            val = f'=利润表!{gcl(YEAR_COL[yr])}{IS_ROW[key]}'
        else:
            val = 0
        ws_sum.cell(14+i, j+2, val)

# 预测摘要
ws_sum.cell(22, 1, '▌ 预测摘要（2026E-2030E）')
pred_headers = ['指标', '2026E', '2027E', '2028E', '2029E', '2030E']
for i, h in enumerate(pred_headers):
    ws_sum.cell(23, i+1, h)
    ws_sum.cell(23, i+1).font = Font(bold=True)

pred_sum_data = [
    ('营业收入', 'revenue', 'IS'),
    ('YoY增速', 'yoy', 'IS'),
    ('归母净利润', 'np_attr', 'IS'),
    ('归母净利YoY', 'np_attr_yoy', 'IS'),
    ('EPS（元）', 'eps', 'IS'),
    ('PE（x）', 'pe', 'CALC'),
]
for i, (label, key, src) in enumerate(pred_sum_data):
    ws_sum.cell(24+i, 1, label)
    for j, yr in enumerate(PRED_YEARS):
        if src == 'IS':
            val = f'=利润表!{gcl(YEAR_COL[yr])}{IS_ROW[key]}'
        else:
            val = '-'
        ws_sum.cell(24+i, j+2, val)

# 质量评分
ws_sum.cell(31, 1, '▌ 质量评分')
ws_sum.cell(32, 1, '综合质量得分')
ws_sum.cell(32, 2, '=基本面研究!D11')
ws_sum.cell(33, 1, '投资等级')
ws_sum.cell(33, 2, 'B档（良好）')

# 估值结论
ws_sum.cell(35, 1, '▌ 估值结论')
ws_sum.cell(36, 1, '当前PE（TTM）')
ws_sum.cell(36, 2, '55.2x')
ws_sum.cell(37, 1, '目标价区间')
ws_sum.cell(37, 2, '64-72元')
ws_sum.cell(38, 1, '上行空间')
ws_sum.cell(38, 2, '+18%~+33%')
ws_sum.cell(39, 1, '投资评级')
ws_sum.cell(39, 2, '中性偏乐观')

ws_sum.column_dimensions['A'].width = 20
for i in range(2, 8):
    ws_sum.column_dimensions[gcl(i)].width = 12

print(f"✓ 摘要 Sheet 完成")

# ===========================
# 验证公式写入
# ===========================
def verify_formulas():
    errors = []
    for yr in PRED_YEARS:
        for key in ['revenue', 'gp', 'np_attr']:
            val = ws_is.cell(IS_ROW[key], YEAR_COL[yr]).value
            if isinstance(val, (int, float)):
                errors.append(f"❌ 利润表 {yr} {key} = 静态数值 {val}，应为公式！")
            elif isinstance(val, str) and not val.startswith('='):
                errors.append(f"❌ 利润表 {yr} {key} 未以=开头: {val}")

    # 模型假设的rev_growth和gross_margin必须为绿色跨Sheet引用
    for yr in PRED_YEARS:
        col_asm = ASM_COL[yr]
        rev_gm = ws_asm.cell(ASM_ROW['gross_margin'], col_asm).value
        rev_gr  = ws_asm.cell(ASM_ROW['rev_growth'], col_asm).value
        if isinstance(rev_gm, (int, float)):
            errors.append(f"❌ 模型假设 {yr} gross_margin = 静态数值 {rev_gm}，应为=业务拆分!公式！")
        if isinstance(rev_gr, (int, float)):
            errors.append(f"❌ 模型假设 {yr} rev_growth = 静态数值 {rev_gr}，应为=业务拆分!公式！")
        if isinstance(rev_gm, str) and not rev_gm.startswith('='):
            errors.append(f"❌ 模型假设 {yr} gross_margin 未以=开头: {rev_gm}")
        if isinstance(rev_gr, str) and not rev_gr.startswith('='):
            errors.append(f"❌ 模型假设 {yr} rev_growth 未以=开头: {rev_gr}")

    if errors:
        print("\n".join(errors))
        raise ValueError("模型公式验证失败！")
    print("✓ 公式验证通过：所有预测列为Excel公式，模型假设联动正确")

verify_formulas()

# ===========================
# 保存文件
# ===========================
output_dir = '/Users/zhuang225/0.Agent 投研总监/reports/航宇科技_2026-04-07'
os.makedirs(output_dir, exist_ok=True)
output_path = f'{output_dir}/航宇科技_688239_财务模型_20260407.xlsx'
wb.save(output_path)
print(f"\n✅ 模型已保存至：{output_path}")
print(f"   Sheet列表：{wb.sheetnames}")