#!/usr/bin/env python3
"""
verify_model.py — 公司研究模型验证脚本
每次构建模型后运行，检测常见的模型问题并给出修复建议。

用法：
    python3 verify_model.py <model_path>
    python3 verify_model.py 阜博集团_03738.HK_研究模型_20260411.xlsx
"""

import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl

PRED_YEARS = ['2026E', '2027E', '2028E', '2029E', '2030E']
ASM_COL = {'2026E': 2, '2027E': 3, '2028E': 4, '2029E': 5, '2030E': 6}

# 行号常量
ASM_ROW = {
    'rev_growth':   5,
    'gross_margin': 8,   # Row7=标题, Row8=综合毛利率
    'sell_rate':    9,
    'admin_rate':   10,
    'rd_rate':     11,
    'finance_cost':12,
    'tax_rate':    13,
    'parent_ratio':14,
    'da':          16,
    'capex':       17,
    'ar_days':     19,
    'inv_days':    20,
    'ap_days':     21,
    'div_payout':  23,
    'shares':      29,
}

IS_ROW = {
    'revenue': 4, 'yoy': 5, 'cogs': 6, 'gp': 7, 'gm': 8,
    'selling': 10, 'admin': 11, 'rd': 12, 'rd_rate': 13, 'finance': 14,
    'op': 16, 'op_margin': 17, 'ebitda': 18, 'ebitda_margin': 19,
    'np': 21, 'np_attr': 22, 'np_attr_yoy': 23, 'np_margin': 24, 'eps': 25,
}

BS_ROW = {
    'cash': 5, 'ar': 6, 'inv': 7, 'other_ca': 8, 'ca': 9,
    'fa': 11, 'other_nca': 12, 'nca': 13, 'ta': 14,
    'std': 17, 'ap': 18, 'other_cl': 19, 'cl': 20,
    'ltd': 22, 'other_ncl': 23, 'ncl': 24, 'tl': 25,
    'parent_eq': 28, 'minority': 29, 'equity': 30,
    'bs_check': 31, 'leverage': 33, 'net_debt_eq': 34,
}

CF_ROW = {
    'oper': 4, 'capex': 5, 'fcf': 6,
    'invest': 8, 'finance_cf': 9,
    'net_chg': 11, 'beg_cash': 12, 'end_cash': 13,
}


class ModelVerifier:
    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path, data_only=False)
        self.errors = []
        self.warnings = []
        self.passed = []

    def _pass(self, msg):
        self.passed.append(msg)

    def _fail(self, msg):
        self.errors.append(msg)

    def _warn(self, msg):
        self.warnings.append(msg)

    def _is_formula(self, v):
        return isinstance(v, str) and v.startswith('=')

    def _detect_year_columns(self, sheet_name, header_rows=None):
        """动态检测各年份所在的列号。尝试多个可能的行号。"""
        if header_rows is None:
            header_rows = [1, 2, 3]
        ws = self.wb[sheet_name]
        for row in header_rows:
            cols = {}
            for col in range(1, 20):
                v = ws.cell(row=row, column=col).value
                if v and str(v) in PRED_YEARS:
                    cols[str(v)] = col
            if cols:
                return cols
        return {}

    def _find_asm_col_refs(self, formula):
        """从公式中提取 模型假设!$X$Y 中的列字母 X。
        专门匹配模型假设Sheet的列引用。"""
        if '模型假设' not in formula:
            return None
        # 匹配 模型假设!$X$数字 或 模型假设!X数字
        m = re.search(r'模型假设![\$]?([A-Z]+)[\$]?(\d+)', formula)
        if m:
            return m.group(1)
        return None

    def run(self):
        print(f"\n{'='*60}")
        print(f"  模型验证报告 — {self.path}")
        print(f"{'='*60}\n")

        self._check_section_header_rows()
        self._check_finance_rate_format()
        self._check_cross_sheet_refs()
        self._check_income_statement_formulas()
        self._check_finance_expense_formula()
        self._check_business_split_predictions()
        self._check_balance_sheet_verification()
        self._check_cash_flow_formulas()
        self._check_summary_sheet_references()
        self._check_asm_col_consistency()

        self._print_report()

    # ──────────────── 检查项 ────────────────

    def _check_section_header_rows(self):
        """检查#1: 分节标题行应为空（不能有公式）"""
        print("[1/9] 检查分节标题行...")
        ws = self.wb['模型假设']
        # Row7 = ▌B.盈利能力（分节标题），B-F列应为空
        row7_has_formula = False
        for col in range(2, 7):  # B-F列
            v = ws.cell(row=7, column=col).value
            if self._is_formula(v):
                self._fail(f"模型假设!{gcl(col)}7 分节标题行有公式: {str(v)[:60]}")
                row7_has_formula = True
        if not row7_has_formula:
            self._pass("模型假设!B7-F7 分节标题行为空")

        # 检查其他分节标题行
        header_rows = [3, 6, 15, 18, 22, 24]
        for row in header_rows:
            v = ws.cell(row=row, column=2).value
            if self._is_formula(v):
                self._fail(f"模型假设!B{row} 分节标题行有公式: {str(v)[:50]}")
                return
        self._pass("其他分节标题行无公式")

    def _check_finance_rate_format(self):
        """检查#2: 财务费用率应为百分比格式（不是绝对值）"""
        print("[2/9] 检查财务费用率格式...")
        ws = self.wb['模型假设']
        label = ws.cell(row=12, column=1).value
        if label and '率' not in str(label):
            self._fail(f"模型假设!A12 标签应为'财务费用率'，当前: '{label}'")
            return
        for col in range(2, 7):
            v = ws.cell(row=12, column=col).value
            if v and isinstance(v, (int, float)) and v > 1:
                self._fail(f"模型假设!{gcl(col)}12 值={v}，应为百分比小数（如0.004），不是绝对值")
                return
        self._pass("模型假设!A12 财务费用率格式正确（百分比）")

    def _check_cross_sheet_refs(self):
        """检查#3: 跨Sheet引用公式正确性"""
        print("[3/9] 检查跨Sheet引用...")
        ws = self.wb['模型假设']
        # Row5（营收增长率）应引用 业务拆分
        rev_ref_ok = True
        for yr, col in ASM_COL.items():
            v = ws.cell(row=5, column=col).value
            if not v or not self._is_formula(v) or '业务拆分' not in v:
                self._fail(f"模型假设!{gcl(col)}5 缺少业务拆分跨Sheet引用 ({yr})")
                rev_ref_ok = False
                break
        if rev_ref_ok:
            self._pass("模型假设!B5-F5 营收增长率→业务拆分引用正确")

        # Row8（综合毛利率）应引用 业务拆分
        gm_ref_ok = True
        for yr, col in ASM_COL.items():
            v = ws.cell(row=8, column=col).value
            if not v or not self._is_formula(v) or '业务拆分' not in v:
                self._fail(f"模型假设!{gcl(col)}8 缺少业务拆分跨Sheet引用 ({yr})")
                gm_ref_ok = False
                break
        if gm_ref_ok:
            self._pass("模型假设!B8-F8 毛利率→业务拆分引用正确")

    def _check_income_statement_formulas(self):
        """检查#4: 利润表预测列公式完整性"""
        print("[4/9] 检查利润表预测列公式...")
        ws = self.wb['利润表']
        data_cols = self._detect_year_columns('利润表', header_rows=[1, 2, 3])

        if not data_cols:
            self._fail("利润表中未检测到预测年份列（尝试Row1/2/3）")
            return

        required_formula_rows = [
            IS_ROW['revenue'], IS_ROW['cogs'], IS_ROW['gp'],
            IS_ROW['selling'], IS_ROW['admin'], IS_ROW['rd'],
            IS_ROW['finance'], IS_ROW['op'], IS_ROW['ebitda'],
            IS_ROW['np'], IS_ROW['np_attr'], IS_ROW['eps'],
        ]
        all_ok = True
        for yr in PRED_YEARS:
            if yr not in data_cols:
                self._warn(f"利润表缺少 {yr} 列")
                continue
            col = data_cols[yr]
            for row in required_formula_rows:
                v = ws.cell(row=row, column=col).value
                if not v or not self._is_formula(v):
                    self._fail(f"利润表!{gcl(col)}{row} ({yr}) 应为公式，当前: {v}")
                    all_ok = False
        if all_ok:
            self._pass("利润表预测列（2026E-2030E）公式完整")

    def _check_finance_expense_formula(self):
        """检查#5: 财务费用公式应为 营收×费率，不是绝对值"""
        print("[5/9] 检查财务费用公式...")
        ws = self.wb['利润表']
        data_cols = self._detect_year_columns('利润表', header_rows=[1, 2, 3])
        if not data_cols or '2026E' not in data_cols:
            self._fail("利润表中未检测到2026E列")
            return
        col = data_cols['2026E']
        v = ws.cell(row=IS_ROW['finance'], column=col).value
        if v and self._is_formula(v):
            if '模型假设' in v:
                self._pass("利润表财务费用公式引用模型假设")
            else:
                self._fail(f"利润表财务费用公式未引用模型假设: {str(v)[:60]}")
        else:
            self._fail(f"利润表!{gcl(col)}{IS_ROW['finance']} 无公式")

    def _check_business_split_predictions(self):
        """检查#6: 业务拆分预测列格式检查
        - 增速/毛利率的原始分段行应为手动值
        - 汇总行（总收入/总毛利/综合毛利率/营收增速）允许公式
        - 收入行允许公式（growth-rate-driven 模式）
        """
        print("[6/9] 检查业务拆分预测列...")
        ws = self.wb['业务拆分']
        data_cols = self._detect_year_columns('业务拆分', header_rows=[1, 2, 3])
        if not data_cols:
            self._warn("业务拆分中未检测到预测年份列")
            return

        issues = []
        for yr, col in data_cols.items():
            for row in range(2, 20):
                v = ws.cell(row=row, column=col).value
                if not self._is_formula(v):
                    continue
                label = str(ws.cell(row=row, column=1).value or '')

                # 汇总行允许公式
                if '汇总' in label or '综合' in label:
                    continue
                # 营收增速汇总行允许公式
                if label.startswith('营收增速'):
                    continue

                # 收入行：允许公式（growth-rate-driven 模式）
                if '收入' in label:
                    continue

                # 毛利额行允许公式
                if '毛利额' in label:
                    continue

                # 其他行不应该有公式
                issues.append(f"业务拆分!{gcl(col)}{row} ({label}) 预测列有公式: {str(v)[:50]}")

        if not issues:
            self._pass("业务拆分预测列格式正确（增速/毛利率=手动值，收入/毛利额/汇总=公式）")
        else:
            for issue in issues:
                self._warn(issue)

    def _check_balance_sheet_verification(self):
        """检查#7: 资产负债表平衡验证"""
        print("[7/9] 检查资产负债表验证...")
        ws = self.wb['资产负债表']
        data_cols = self._detect_year_columns('资产负债表', header_rows=[1, 2, 3])
        if not data_cols:
            self._warn("资产负债表中未检测到预测年份列")
            return
        all_ok = True
        for yr in PRED_YEARS:
            if yr not in data_cols:
                continue
            col = data_cols[yr]
            v = ws.cell(row=BS_ROW['bs_check'], column=col).value
            if not v or not self._is_formula(v):
                self._fail(f"资产负债表!{gcl(col)}{BS_ROW['bs_check']} 缺少验证公式 ({yr})")
                all_ok = False
        if all_ok:
            self._pass("资产负债表验证公式完整")

    def _check_cash_flow_formulas(self):
        """检查#8: 现金流量表期末现金公式"""
        print("[8/9] 检查现金流量表公式...")
        ws = self.wb['现金流量表']
        data_cols = self._detect_year_columns('现金流量表', header_rows=[1, 2, 3])
        if not data_cols:
            self._warn("现金流量表中未检测到预测年份列")
            return
        all_ok = True
        for yr in PRED_YEARS:
            if yr not in data_cols:
                continue
            col = data_cols[yr]
            v = ws.cell(row=CF_ROW['end_cash'], column=col).value
            if not v or not self._is_formula(v):
                self._fail(f"现金流量表!{gcl(col)}{CF_ROW['end_cash']} 缺少期末现金公式 ({yr})")
                all_ok = False
        if all_ok:
            self._pass("现金流量表期末现金公式完整")

    def _check_summary_sheet_references(self):
        """检查#9: 摘要页预测列跨Sheet引用"""
        print("[9/9] 检查摘要页跨Sheet引用...")
        ws = self.wb['摘要']

        # 摘要页布局可能不标准，尝试多种行号
        data_cols = self._detect_year_columns('摘要', header_rows=list(range(1, 20)))
        if not data_cols:
            self._warn("摘要页未检测到预测年份列（可能布局非标准）")
            return

        # 查找关键指标行
        summary_keywords = ['营业收入', '归母净利', '毛利率', 'ROE']
        found_rows = {}
        for keyword in summary_keywords:
            for row in range(4, 25):
                label = str(ws.cell(row=row, column=1).value or '')
                if keyword in label:
                    found_rows[keyword] = row
                    break

        if not found_rows:
            self._warn("摘要页未找到关键指标行")
            return

        # 检查预测列是否有公式引用
        missing_refs = []
        for yr in ['2026E', '2027E']:
            if yr not in data_cols:
                continue
            col = data_cols[yr]
            for keyword, row in found_rows.items():
                v = ws.cell(row=row, column=col).value
                if not v or not self._is_formula(v):
                    missing_refs.append(f"摘要!{gcl(col)}{row} ({keyword}) 缺少公式 ({yr})")

        if not missing_refs:
            self._pass("摘要页预测列跨Sheet引用完整")
        else:
            for m in missing_refs:
                self._warn(m)

    def _check_asm_col_consistency(self):
        """检查#10: 模型假设列引用一致性
        每个预测年的公式应引用模型假设中对应的列，不能全部引用同一列。
        """
        print("[10/9] 检查模型假设列引用一致性...")

        # 方法1：检查利润表 revenue 行（Row4）的公式引用
        is_cols = self._detect_year_columns('利润表', header_rows=[1, 2, 3])
        if not is_cols:
            self._warn("利润表未检测到年份列，无法检查模型假设列引用一致性")
            return

        ws_is = self.wb['利润表']
        revenue_row = IS_ROW['revenue']

        # ASM列映射：年份 → 模型假设中的列字母
        ASM_LETTER = {yr: gcl(col) for yr, col in ASM_COL.items()}
        expected_refs = {}
        actual_refs = {}

        for yr in PRED_YEARS:
            if yr not in is_cols:
                continue
            col = is_cols[yr]
            v = ws_is.cell(row=revenue_row, column=col).value
            if v and self._is_formula(v):
                # 提取 模型假设!$X$数字 中的列字母
                m = re.search(r'模型假设![\$]?([A-Z]+)[\$]?(\d+)', v)
                if m:
                    actual_refs[yr] = m.group(1)
                    expected_refs[yr] = ASM_LETTER[yr]

        if not actual_refs:
            self._warn("未能从利润表公式中提取模型假设列引用")
            return

        # 检查是否有引用错误的
        mismatches = []
        for yr in actual_refs:
            if actual_refs[yr] != expected_refs.get(yr):
                mismatches.append(f"{yr}: 引用${actual_refs[yr]}$, 应为${expected_refs.get(yr, '?')}$")

        if mismatches:
            for m in mismatches:
                self._fail(f"利润表列引用错误 — {m}")
        else:
            self._pass(f"利润表预测列引用正确: {actual_refs}")

    # ──────────────── 报告输出 ────────────────

    def _print_report(self):
        print(f"\n{'─'*60}")
        print(f"  通过: {len(self.passed)} 项")
        for p in self.passed:
            print(f"  ✅ {p}")

        if self.warnings:
            print(f"\n  ⚠️  警告: {len(self.warnings)} 项")
            for w in self.warnings:
                print(f"  ⚠️  {w}")

        if self.errors:
            print(f"\n  ❌ 错误: {len(self.errors)} 项")
            for e in self.errors:
                print(f"  ❌ {e}")

        print(f"\n{'─'*60}")
        total = len(self.passed) + len(self.warnings) + len(self.errors)
        if not self.errors and not self.warnings:
            print(f"  ✅ 全部通过 ({total} 项) — 模型质量优秀")
        elif not self.errors:
            print(f"  ⚠️  {len(self.warnings)} 项警告，建议检查 ({total} 项)")
        else:
            print(f"  ❌ 发现 {len(self.errors)} 个问题，需要修复 ({total} 项)")
        print(f"{'='*60}\n")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python3 verify_model.py <模型文件路径>")
        print("示例: python3 verify_model.py 阜博集团_03738.HK_研究模型_20260411.xlsx")
        sys.exit(1)

    path = sys.argv[1]
    verifier = ModelVerifier(path)
    verifier.run()
